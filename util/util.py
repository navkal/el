# Copyright 2019 Energize Andover.  All rights reserved.

import os
import sqlite3
import sqlalchemy
import pandas as pd
import openpyxl
import re
import string
import datetime
import warnings

import time
START_TIME = time.time()
print( 'Starting at', time.strftime( '%H:%M:%S', time.localtime( START_TIME ) ) )


ELECTION_TYPES = \
{
    'STATE_ELECTION': 'STATE ELECTION',
    'SPECIAL_STATE': 'SPECIAL STATE',
    'PRESIDENTIAL_PRIMARY': 'PRESIDENTIAL PRIMARY',
    'STATE_PRIMARY': 'STATE PRIMARY',
    'SPECIAL_STATE_PRIMARY': 'SPECIAL STATE PRIMARY',
    'LOCAL_ELECTION': 'LOCAL ELECTION',
    'LOCAL_TOWN_MEETING': 'LOCAL TOWN MEETING',
}

ID = 'id'

R = 'R'
D = 'D'
U = 'U'

ABSENT = '-'
VOTED_BOTH = 'voted_both'
CHANGED_AFFILIATION = 'changed_affiliation'
PARTY_VOTED_HISTORY = 'party_voted_history'
LIKELY_DEM_SCORE = 'likely_dem_score'
PARTY_PREFERENCE_SCORE = 'party_preference_score'
LEGACY_PREFERENCE_SCORE = 'legacy_preference_score'
LEGACY_DEM_SCORE = 'legacy_dem_score'
VOTER_ENGAGEMENT_SCORE = 'voter_engagement_score'
TOWN_MEETINGS_ATTENDED = 'town_meetings_attended'
PRIMARY_ELECTIONS_VOTED = 'primary_elections_voted'
LOCAL_ELECTIONS_VOTED = 'local_elections_voted'
STATE_ELECTIONS_VOTED = 'state_elections_voted'
SPECIAL_ELECTIONS_VOTED = 'special_elections_voted'
EARLY_VOTES = 'early_votes'
VOTED = 'voted'

PARTISAN_SCORE = 'partisan_score'
RECENT_LOCAL_ELECTION_COUNT = 3
RECENT_LOCAL_ELECTIONS_VOTED = 'recent_local_elections_voted'

ELECTION_YEAR = 'election_year'
TURNOUT = 'turnout'
TURNOUT_FACTOR = 'turnout_factor'
RECENCY_FACTOR = 'recency_factor'
SCORE = 'score'

YES = 'Y'
NO = 'N'

IS_HOMEOWNER = 'is_homeowner'
IS_FAMILY = 'is_family'
LEGACY_AGE_MAX = 30
VOTER_AGE_MIN = 18

CONVENIENCE = 'CONVENIENCE'
FAMILY = 'FAMILY'
NAL_IGNORE = [ CONVENIENCE, FAMILY ]

RESIDENT_COUNT = 'resident_count'
VOTER_COUNT = 'voter_count'
LIKELY_DEM_COUNT = 'likely_dem_count'
LIKELY_REPUB_COUNT = 'likely_repub_count'
LOCAL_DEM_VOTER_COUNT = 'local_dem_voter_count'
LOCAL_REPUB_VOTER_COUNT = 'local_repub_voter_count'

MEAN_LIKELY_DEM_SCORE = 'mean_likely_dem_score'
MEAN_PARTY_PREFERENCE_SCORE = 'mean_party_preference_score'
MEAN_VOTER_ENGAGEMENT_SCORE = 'mean_voter_engagement_score'
MEAN_TOTAL_ASSESSED_VALUE = 'mean_total_assessed_value'
MEAN_LIKELY_DEM_VOTER_ENGAGEMENT_SCORE = 'mean_likely_dem_voter_engagement_score'
MEAN_LIKELY_REPUB_VOTER_ENGAGEMENT_SCORE = 'mean_likely_repub_voter_engagement_score'

GAL_PER_CU_FT = 7.48052

METER_STATUS = 'meter_status'
METER_WRAP = 'wrap'
METER_NORMAL = 'normal'
METER_ANOMALY = 'anomaly'



# Mappings to reconcile column names

CENSUS_YEAR = 'census_year'
FIRST_NAME = 'first_name'
MIDDLE_NAME = 'middle_name'
LAST_NAME = 'last_name'
TITLE = 'title'
DATE_OF_BIRTH = 'date_of_birth'
ACCOUNT_NUMBER = 'account_number'
AGE = 'age'

PARTY_AFFILIATION = 'party_affiliation'
RESIDENT_ID = 'resident_id'
ELECTION_TYPE = 'election_type'
VOTER_STATUS = 'voter_status'
OCCUPATION = 'occupation'
HEAD_OF_HOUSEHOLD = 'head_of_household'
FAMILY_GROUP_CODE = 'family_group_code'
VETERAN = 'veteran'

STREET_NUMBER = '{0}_street_number'
STREET_NUMBER_SUFFIX = '{0}_street_number_suffix'
STREET_NAME = '{0}_street_name'
APARTMENT_NUMBER = '{0}_apartment_number'
ZIP_CODE = '{0}_zip_code'

NORMALIZED_ADDRESS = 'normalized_address'
NORMALIZED_STREET_NUMBER = 'street_number'
NORMALIZED_STREET_NAME = 'street_name'
NORMALIZED_OCCUPANCY = 'occupancy'
NORMALIZED_ADDITIONAL_INFO = 'additional_address_info'

LEFT_ADDR_FULL = 'left_addr_full'
LEFT_ADDR_TRUNC = 'left_addr_trunc'
LEFT_ADDR_EDIT = 'left_addr_edit'
LEFT_ADDR_STRIP = 'left_addr_strip'
RIGHT_ADDR_TRUNC = 'right_addr_trunc'
RIGHT_ADDR_EDIT = 'right_addr_edit'
RIGHT_ADDR_STRIP = 'right_addr_strip'

# Confidence of address matching results
CONFIDENCE = 'confidence'
CONFIDENCE_HIGH = 'High'
CONFIDENCE_MEDIUM = 'Medium'
CONFIDENCE_LOW = 'Low'

ADDRESS = 'address'
ADDR_STREET_NUMBER = STREET_NUMBER.format( ADDRESS )
ADDR_STREET_NAME = STREET_NAME.format( ADDRESS )

LOCATION_ADDRESS = 'location_address'
LADDR_STREET_NUMBER = STREET_NUMBER.format( LOCATION_ADDRESS )
LADDR_ALT_STREET_NUMBER = STREET_NUMBER.format( LOCATION_ADDRESS + '_alternate' )
LADDR_STREET_NAME = STREET_NAME.format( LOCATION_ADDRESS )
LADDR_CITY = LOCATION_ADDRESS + '_city'
LADDR_CONDO_UNIT = LOCATION_ADDRESS + '_condo_unit'
LADDR_CONDO_COMPLEX = LOCATION_ADDRESS + '_condo_complex'

RESIDENTIAL_ADDRESS = 'residential_address'
RADDR_STREET_NUMBER = STREET_NUMBER.format( RESIDENTIAL_ADDRESS )
RADDR_STREET_NUMBER_SUFFIX = STREET_NUMBER_SUFFIX.format( RESIDENTIAL_ADDRESS )
RADDR_STREET_NAME = STREET_NAME.format( RESIDENTIAL_ADDRESS )
RADDR_APARTMENT_NUMBER = APARTMENT_NUMBER.format( RESIDENTIAL_ADDRESS )
RADDR_ZIP_CODE = ZIP_CODE.format( RESIDENTIAL_ADDRESS )
PRECINCT_NUMBER = 'precinct_number'

MAILING_ADDRESS = 'mailing_address'
MADDR_APARTMENT_NUMBER = APARTMENT_NUMBER.format( MAILING_ADDRESS )
MADDR_CITY = MAILING_ADDRESS + '_city'
MADDR_STATE = MAILING_ADDRESS + '_state'
MADDR_STREET = MAILING_ADDRESS + '_street'
MADDR_CSZ = MAILING_ADDRESS + '_csz'
MADDR_ZIP_CODE = ZIP_CODE.format( MAILING_ADDRESS )
MADDR_LINE = MAILING_ADDRESS + '_line_{0}'

TOWN_CODE = 'town_code'
TOWN_NAME = 'town_name'
TOWN_INDICATOR = 'town_indicator'

WARD_NUMBER = 'ward_number'

ELECTION_DATE = 'election_date'
PARTY_VOTED = 'party_voted'

BALLOT_MAILED_DISPOSITION = 'ballot_{0}_mailed_disposition'
BALLOT_RECEIVED_BY = 'ballot_{0}_received_by'
BALLOT_REJECTION_REASON = 'ballot_{0}_rejection_reason'
BALLOT_SENT_BY = 'ballot_{0}_sent_by'
BALLOT_MAILED_DATE = 'ballot_{0}_mailed_date'
BALLOT_RECEIVED_DATE = 'ballot_{0}_received_date'

PARCEL_ID = 'parcel_id'
PHONE = 'phone'
METER_NUMBER = 'meter_number'
SERVICE_ID = 'service_id'
DESCRIPTION = 'description'
_DESC = '_' + DESCRIPTION
CURRENT_DATE = 'current_date'
CURRENT_READING = 'current_reading'
PRIOR_DATE = 'prior_date'
PRIOR_READING = 'prior_reading'
TRANSACTION_DATE = 'transaction_date'
TRANSACTION_TYPE = 'transaction_type'
UNITS = 'units'
SERVICE_TYPE = 'service_type'
SERVICE = 'service'
CU_FT = 'cu_ft'
ELAPSED_DAYS = 'elapsed_days'
CU_FT_PER_DAY = 'cu_ft_per_day'
GAL_PER_DAY = 'gal_per_day'

OWNER_1_NAME = 'owner_1_name'
OWNER_2_NAME = 'owner_2_name'
OWNER_3_NAME = 'owner_3_name'
OWNER_ADDRESS = 'owner_address'
OWNER_ZIP = 'owner_zip'
SITE_ADDRESS = 'site_address'

SALE_DATE = 'sale_date'
LEGAL_REFERENCE_SALE_DATE = 'legal_reference_sale_date'
NAL_DESCRIPTION = 'nal_description'
PREVIOUS_LEGAL_REFERENCE_SALE_DATE = 'previous_legal_reference_sale_date'
PREVIOUS_NAL_DESCRIPTION = 'previous_nal_description'
GRANTOR = 'grantor'
PREVIOUS_GRANTOR = 'previous_grantor'
OWNER_OCCUPIED = 'owner_occupied'
ZONING_CODE_1 = 'zoning_code_1'
TOTAL_ASSESSED_VALUE = 'total_assessed_value'
ZONE = 'zone'

DEPARTMENT = 'department'
POSITION = 'position'

GENDER = 'gender'

OFFICE_OR_CANDIDATE = 'office_or_candidate'
OFFICE = 'office'
CANDIDATE = 'candidate'
PRECINCT = 'precinct_'
PRECINCT_1 = PRECINCT + '1'
PRECINCT_2 = PRECINCT + '2'
PRECINCT_3 = PRECINCT + '3'
PRECINCT_4 = PRECINCT + '4'
PRECINCT_5 = PRECINCT + '5'
PRECINCT_6 = PRECINCT + '6'
PRECINCT_7 = PRECINCT + '7'
PRECINCT_8 = PRECINCT + '8'
PRECINCT_9 = PRECINCT + '9'
TOTAL = 'total'

PIN = 'PIN'
PERMITFOR = 'PermitFor'
PERMIT_FOR = 'permit_for'
DATE_ISSUED = 'date_issued'
TOTALFEE = 'TotalFee'
PROJECTCOST = 'ProjectCost'
COST = 'Cost'


ZIP = 'zip_code'
ZIP_CODES = 'zip_codes'
YEAR = 'year'
COUNTY = 'county'
SECTOR = 'sector'

SECTOR_RES_AND_LOW = 'Residential & Low-Income'
SECTOR_COM_AND_IND = 'Commercial & Industrial'
SECTOR_TOTAL = 'Total'

COMBINED_EES_IN = 'combined_ees_in_$'
COMBINED_INCENTIVES_OUT = 'combined_incentives_out_$'
COMBINED_EES_MINUS_INCENTIVES = 'combined_ees_minus_incentives_$'
MWH_SAVED_AS_PCT_OF_USED = 'mwh_saved_as_%_of_used'
THERMS_SAVED_AS_PCT_OF_USED = 'therms_saved_as_%_of_used'

JAN = 'jan'
FEB = 'feb'
MAR = 'mar'
APR = 'apr'
MAY = 'may'
JUN = 'jun'
JUL = 'jul'
AUG = 'aug'
SEP = 'sep'
OCT = 'oct'
NOV = 'nov'
DEC = 'dec'

ANNUAL = 'annual'
ELECTRIC_USAGE = '_electric_usage_mwh'
GAS_USAGE = '_gas_usage_therms'
ELECTRIC_SAVINGS = '_electric_savings_mwh'
GAS_SAVINGS = '_gas_savings_therms'
ANNUAL_ELECTRIC_USAGE = ANNUAL + ELECTRIC_USAGE
ANNUAL_GAS_USAGE = ANNUAL + GAS_USAGE
ANNUAL_ELECTRIC_SAVINGS = ANNUAL + ELECTRIC_SAVINGS
ANNUAL_GAS_SAVINGS = ANNUAL + GAS_SAVINGS
ELECTRIC_INCENTIVES = 'electric_incentives_$'
GAS_INCENTIVES = 'gas_incentives_$'
POPULATION = 'population'
TRACT_POPULATION = 'tract_population'
PCT_LOW_INCOME = 'pct_low_income'
PCT_ENERGY_BURDENED = 'pct_energy_burdened'
RESIDENTIAL_DOL_PER_MWH = 'residential_$_per_mwh'
DISCOUNT_DOL_PER_MWH = 'discount_$_per_mwh'
COMMERCIAL_DOL_PER_MWH = 'commercial_$_per_mwh'
RESIDENTIAL_DOL_PER_THERM = 'residential_$_per_therm'
COMMERCIAL_DOL_PER_THERM = 'commercial_$_per_therm'
INCENTIVES_PER_SAVED_MWH = 'incentives_per_saved_mwh_$'
INCENTIVES_PER_SAVED_THERM = 'incentives_per_saved_therm_$'

TOTAL_ANNUAL_ELECTRIC_USAGE = 'tot_' + ANNUAL_ELECTRIC_USAGE
TOTAL_ANNUAL_ELECTRIC_SAVINGS = 'tot_' + ANNUAL_ELECTRIC_SAVINGS
MEDIAN_INCENTIVES_PER_SAVED_MWH = 'median_' + INCENTIVES_PER_SAVED_MWH
AVG_INCENTIVES_PER_SAVED_MWH = 'avg_' + INCENTIVES_PER_SAVED_MWH
STD_INCENTIVES_PER_SAVED_MWH = 'std_' + INCENTIVES_PER_SAVED_MWH

TOTAL_ANNUAL_GAS_USAGE = 'tot_' + ANNUAL_GAS_USAGE
TOTAL_ANNUAL_GAS_SAVINGS = 'tot_' + ANNUAL_GAS_SAVINGS
MEDIAN_INCENTIVES_PER_SAVED_THERM = 'median_' + INCENTIVES_PER_SAVED_THERM
AVG_INCENTIVES_PER_SAVED_THERM = 'avg_' + INCENTIVES_PER_SAVED_THERM
STD_INCENTIVES_PER_SAVED_THERM = 'std_' + INCENTIVES_PER_SAVED_THERM

ELECTRIC_UTILITY = 'electric_utility'
ELECTRIC_UTILITY_URL = 'electric_utility_url'

GAS_UTILITY = 'gas_utility'
GAS_UTILITY_1 = GAS_UTILITY + '_1'
GAS_UTILITY_2 = GAS_UTILITY + '_2'
GAS_UTILITY_URL_1 = GAS_UTILITY + '_url_1'
GAS_UTILITY_URL_2 = GAS_UTILITY + '_url_2'

RESIDENTIAL_RATE = 'residential_rate'
RESIDENTIAL_R1_RATE = 'residential_r1_rate'
RESIDENTIAL_R2_RATE = 'residential_r2_rate'
COMMERCIAL_RATE = 'commercial_rate'

VISION_ID = 'vision_id'
VISION_LINK = 'vision_link'
GIS_ID = 'gis_id'
OWNER_NAME = 'owner_name'
LOCATION = 'location'
YEAR_BUILT = 'year_built'
EFFECTIVE_YEAR_BUILT = 'effective_year_built'
TOTAL_ACRES = 'total_acres'
BUILDING_COUNT = 'building_count'
PARCEL_COUNT = 'parcel_count'
GROSS_BUILDING_AREA = 'gross_building_area'
LIVING_AREA = 'living_area'
YEAR_REMODELED = 'year_remodeled'
BASEMENT = 'basement'
BATHROOM_STYLE = 'bathroom_style'
FIREPLACES = 'number_of_fireplaces'
HEATING_TYPE = 'heating_type'
KITCHEN_STYLE = 'kitchen_style'
FULL_BATHS = 'number_of_full_baths'
HALF_BATHS = 'number_of_half_baths'
KITCHENS = 'number_of_kitchens'
BATHS = 'number_of_baths'
BEDROOOMS = 'number_of_bedrooms'
ROOMS = 'number_of_rooms'
ROOF_TYPE = 'roof_type'
SIDING_TYPE = 'siding_type'
ROOF_COVER = 'roof_cover'
ROOF_STRUCTURE = 'roof_structure'
HEATING_FUEL = 'heating_fuel'
LAND_USE_CODE = 'primary_land_use_code'
SALE_PRICE = 'sale_price'
STORY_HEIGHT = 'story_height'
AC_TYPE = 'ac_type'
HEAT_AC = 'heat_ac'
_CODE = '_code'
BOOK = 'book'
PAGE = 'page'
STATUS = 'status'
BUILDING_CONDITION = 'building_condition'
STYLE = 'style'
WORK_DESCRIPTION = 'work_description'
PERMIT_NUMBER = 'permit_number'
CONTRACTOR_NAME = 'contractor_name'
CONTRACTOR_NAME_1 = ( CONTRACTOR_NAME + '_{}' ).format( 1 )
CONTRACTOR_NAME_2 = ( CONTRACTOR_NAME + '_{}' ).format( 2 )
CONTRACTOR_NAME_3 = ( CONTRACTOR_NAME + '_{}' ).format( 3 )
DATE = 'date'
APPLICATION_DATE = 'application_date'
OCCUPANCY_TYPE = 'occupancy_type'
RENTAL_LIVING_UNITS = 'rental_living_units'
OCCUPANCY_HOUSEHOLDS = 'occupancy_households'
LICENSE_NUMBER = 'license_number'
EXTERIOR_WALL_TYPE = 'exterior_wall_type'
BUILDING_GRADE = 'building_grade'
GRADE = 'grade'
FIRST_FLOOR_USE = 'first_floor_use'
RESIDENTIAL_UNITS = 'residential_units'
TOTAL_OCCUPANCY = 'total_occupancy'
TOTAL_BATHS = 'total_baths'
TOTAL_KITCHENS = 'total_kitchens'
TOTAL_AREA = 'total_area'

EMPLOYEES = 'number_of_employees'
APPLICANT = 'applicant'
EMAIL = 'email'
APPLICANT_ADDRESS = 'applicant_address'
BUSINESS_MANAGER = 'business_manager'
BUSINESS_NAME = 'business_name'
LICENSE_SUBTYPE = 'license_subtype'
LICENSE_TYPE = 'license_type'
CLOSED_DATE = 'closed_date'

IS_RESIDENTIAL = 'is_residential'
OWNER_IS_LOCAL = 'owner_is_local'

OPENED = 'opened'

_LINK = '_link'
FILE_NUMBER = 'file_number'
FILE_NUMBER_LINK = 'file_number' + _LINK
PERMIT_NUMBER_LINK = PERMIT_NUMBER + _LINK
PERMIT_TYPE = 'permit_type'
TOTAL_FEE = 'total_fee_$'
PROJECT_COST = 'project_cost_$'
TOTAL_PROJECT_COST = 'total_' + PROJECT_COST
ELECTRICAL_PERMIT_NUMBER = 'electrical_' + PERMIT_NUMBER
INSPECTION_NOTES = 'inspection_notes'
PERMIT_SUBTYPE = 'permit_subtype'
PROJECT_COUNT = 'project_count'
PROJECT_TYPE = 'project_type'
PERMIT_REGEX = 'permit_regex'

WATTS_PER_MODULE = 'watts_per_module'
MODULES ='number_of_solar_pv_modules'
PERMIT_STATUS = 'permit_status'
KW_DC = 'kw_dc'
INTERACTIVE_INVERTERS = 'number_of_interactive_inverters'
PROPERTY_USE_GROUP = 'property_use_group'
POWER_RATINGS_EACH = 'power_ratings_each'
APPROVAL_DATE = 'approval_date'
EXPIRATION_DATE = 'expiration_date'
LAST_INSPECTION_DATE = 'last_inspection_date'
INSPECTION_TYPE = 'inspection_type'
ASSIGNED_TO = 'assigned_to'
DATE_DUE_FOR_INSPECTION = 'date_due_for_inspection'
INSPECTION_STATUS = 'inspection_status'
PROPERTY_MANAGER = 'property_manager'

APPLIANCE = 'appliance'

ENTIRE_ROOF = 'entire_roof'
STRIPPING = 'stripping'
ROOF_LAYERS = 'roof_layers'
SUBMITTAL_DOCUMENTS = 'submittal_documents'

USAGE_END_NULL = 'usage_end_null'
ONE = 'one'

CUSTOMER_NAME = 'customer_name'
UTILITY_ACCOUNT = 'utility_account'
EXTERNAL_SUPPLIER = 'external_supplier'
ISO_ZONE = 'iso_zone'
MEI_TRACKING = 'mei_tracking'
STREET = 'street'
CITY = 'city'
STATE = 'state'
COST_OR_USE = 'cost_or_use'

CONTRACT = 'contract'
METHOD = 'method'
PERIOD = 'period'
VENDOR = 'vendor'
VENDOR_NUMBER = 'vendor_number'
VENDOR_NAME = 'vendor_name'
AMOUNT = 'amount_$'
ADDR_1 = 'address_1'
ADDR_2 = 'address_2'
CONTACT_NAME = 'contact_name'
CONTACT_EMAIL = 'contact_email'
JOB_NUMBER = 'job_number'

# Name fragments for Lawrence parcel summary
OIL_ = 'oil_'
GAS_ = 'gas_'
ELEC_ = 'elec_'
KITCHENS_ = 'kitchens_'
BATHS_ = 'baths_'
OCCUPANCY_ = 'occupancy_'

MBLU = 'mblu' # Stands for Map, Block, Lot, Unit

CONSISTENT_COLUMN_NAMES = \
{
    'Assessment': \
    {
        'AccountNumber': ACCOUNT_NUMBER,
        'ParcelID': PARCEL_ID,
        'UserAccount': 'user_account',
        'Street.': LADDR_STREET_NUMBER,
        'AltStreet.': LADDR_ALT_STREET_NUMBER,
        'StreetName': LADDR_STREET_NAME,
        'LocCity': LADDR_CITY,
        'CondoUnit': LADDR_CONDO_UNIT,
        'CondoComplex': LADDR_CONDO_COMPLEX,
        'Owner1': OWNER_1_NAME,
        'Owner2': OWNER_2_NAME,
        'Owner3': OWNER_3_NAME,
        'BillingAddress': MADDR_LINE.format( 1 ),
        'BillingAddress2': MADDR_LINE.format( 2 ),
        'City': MADDR_CITY,
        'State': MADDR_STATE,
        'Zip': MADDR_ZIP_CODE,
        'OwnerOccupied': OWNER_OCCUPIED,
        'Zone1': ZONING_CODE_1,
        'FloodHazard': 'flood_hazard_zone',
        'Census': 'census_code',
        'Utility1': 'utility_code_1',
        'Utility2': 'utility_code_2',
        'Utility3': 'utility_code_3',
        'Traffic': 'traffic_code',
        'LUC': LAND_USE_CODE,
        'YearBuilt': YEAR_BUILT,
        'EffYr': EFFECTIVE_YEAR_BUILT,
        'TotalAcres': TOTAL_ACRES,
        'GrossBldArea': GROSS_BUILDING_AREA,
        'BuildType': 'building_type_1',
        'BuildType2': 'building_type_2',
        'BuildType3': 'building_type_3',
        'FinArea': 'finished_building_area',
        'NumofBuilding': 'number_of_buildings',
        'StoryHeight': STORY_HEIGHT,
        'RentalLivUnits': RENTAL_LIVING_UNITS,
        'Rooms': ROOMS,
        'Bedrooms': BEDROOOMS,
        'Fullbaths': FULL_BATHS,
        'HalfBaths': HALF_BATHS,
        'OtherFixtures': 'number_of_other_fixtures',
        'BathRating': 'bathroom_rating',
        'Kitchens': KITCHENS,
        'KitchenRating': 'kitchen_rating',
        'FirePlaces': FIREPLACES,
        'WSFlues': 'number_of_wood_stove_flues',
        'SolarHotWater': 'solar_hot_water',
        'CentralVac': 'central_vacuum',
        'HeatType': HEATING_TYPE,
        'HeatFuel': HEATING_FUEL,
        'PercentAC': 'percent_air_conditioned',
        'BasementArea': 'basement_area',
        'FinBasementArea': 'finished_basement_area',
        'RoofType': ROOF_TYPE,
        'RoofCover': ROOF_COVER,
        'ExtWall': EXTERIOR_WALL_TYPE,
        'IntWall': 'interior_wall_type',
        'AttachedGarage': 'attached_garage_area',
        'DetachedGarage': 'detached_garage_size',
        'BasementGarage': 'number_of_basement_garages',
        'Pool': 'pool',
        'Frame': 'building_frame_type',
        'Floor': 'floor_type',
        'BaseYear': 'base_depreciation_year',
        'Grade': BUILDING_GRADE,
        'Cond': BUILDING_CONDITION,
        'LegalReference': 'legal_reference_number',
        'LegalRefDate': LEGAL_REFERENCE_SALE_DATE,
        'SalePrice': SALE_PRICE,
        'NAL': NAL_DESCRIPTION,
        'Grantor': GRANTOR,
        'PrevLegalReference': 'previous_legal_reference_number',
        'PrevLegalRefDate': PREVIOUS_LEGAL_REFERENCE_SALE_DATE,
        'PrevSalePrice': 'previous_sale_price',
        'PrevNal': PREVIOUS_NAL_DESCRIPTION,
        'PrevGrantor': PREVIOUS_GRANTOR,
        'TotalLandValue': 'total_land_value',
        'TotalYardItemValue': 'total_yard_item_value',
        'TotalBuildingValue': 'total_building_value',
        'TotalValue': TOTAL_ASSESSED_VALUE,
        'LegalDescription': 'legal_description',
        'AdjArea': 'adjusted_area',
        'PercSprink': 'percolation_sprinkler',
        'RevDate': 'revision_date',
        'GisKey': 'gis_key',
    },
    'AssessmentAddendum': \
    {
        'Street.': LADDR_STREET_NUMBER,
        'AltStreet.': LADDR_ALT_STREET_NUMBER,
        'StreetName': LADDR_STREET_NAME,
        'Zone1': ZONING_CODE_1,
    },
    'BuildingPermits': \
    {
        PIN: 'pin',
        'OwnerName': OWNER_NAME,
        PERMITFOR: PERMIT_FOR,
        'DateIssued': DATE_ISSUED,
        'DateIssued/Submit': DATE_ISSUED,
        'DateIssued/': DATE_ISSUED,
        'ParcelID': PARCEL_ID,
        'House#': ADDR_STREET_NUMBER,
        'Street': ADDR_STREET_NAME,
        'OccupancyType': OCCUPANCY_TYPE,
        'Occ.Type': OCCUPANCY_TYPE,
        'Occ': OCCUPANCY_TYPE,
        'OccType': OCCUPANCY_TYPE,
        'Type': OCCUPANCY_TYPE,
        'BuildingType': 'building_type',
        'BldgType': 'building_type',
        'WorkDescription': WORK_DESCRIPTION,
        'ContractorName': CONTRACTOR_NAME,
        PROJECTCOST: PROJECT_COST,
        COST: PROJECT_COST,
        TOTALFEE: TOTAL_FEE,
    },
    'Census': \
    {
        'Census Yr': CENSUS_YEAR,
        'Res. ID': RESIDENT_ID,
        'Last Nm': LAST_NAME,
        'First Nm': FIRST_NAME,
        'Middle Nm': MIDDLE_NAME,
        'D.O.B. (mm/dd/yyyy format)': DATE_OF_BIRTH,
        'Occupation': OCCUPATION,
        'Mail to Code (Y if HOH; N if Not)': HEAD_OF_HOUSEHOLD,
        'Res. - St. #': RADDR_STREET_NUMBER,
        'Res. - St. # Suffix': RADDR_STREET_NUMBER_SUFFIX,
        'Res. - St. Name': RADDR_STREET_NAME,
        'Res. - Apt #': RADDR_APARTMENT_NUMBER,
        'Res. - Zip Code': RADDR_ZIP_CODE,
        'Party': PARTY_AFFILIATION,
        'Ward #': WARD_NUMBER,
        'Precinct #': PRECINCT_NUMBER,
        'Voter Status': VOTER_STATUS,
    },
    'EjCommunities': \
    {
        'TOWN_NAME': TOWN_NAME,
        'COUNTY_NAME': 'county_name',
        'TRACT_NAME': 'tract_name',
        'Census Tract': 'census_tract',
        'Tot_Pop': TRACT_POPULATION,
        'AvgEnergyBurden': 'avg_energy_burden',
        'AvgEnergyBurden_Screen': 'avg_energy_burden_screen',
        'PCT_LowIncome': PCT_LOW_INCOME,
        'PCT_LowIncome_Screen': 'pct_low_income_screen',
        'FE_Comm': 'fe_comm',
        'FE_Comm.1': 'fe_comm_l',
        'EJ_Indicators': 'ej_indicators',
        'EJ_Comm_Screen': 'ej_comm_screen',
        'P_LDPNT_D2': 'p_ldpnt_d2',
        'P_DSLPM_D2': 'p_dslpm_d2',
        'P_CANCR_D2': 'p_cancr_d2',
        'P_RESP_D2': 'p_resp_d2',
        'P_PTRAF_D2': 'p_ptraf_d2',
        'P_PWDIS_D2': 'p_pwdis_d2',
        'P_PNPL_D2': 'p_pnpl_d2',
        'P_PRMP_D2': 'p_prmp_d2',
        'P_PTSDF_D2': 'p_ptsdf_d2',
        'P_OZONE_D2': 'p_ozone_d2',
        'P_PM25_D2': 'p_pm25_d2',
        'P_60_LDPNT_D2': 'p_60_ldpnt_d2',
        'P_60_DSLPM_D2': 'p_60_dslpm_d2',
        'P_60_CANCR_D2': 'p_60_cancr_d2',
        'P_60_RESP_D2': 'p_60_resp_d2',
        'P_60_PTRAF_D2': 'p_60_ptraf_d2',
        'P_60_PWDIS_D2': 'p_60_pwdis_d2',
        'P_60_PNPL_D2': 'p_60_pnpl_d2',
        'P_60_PRMP_D2': 'p_60_prmp_d2',
        'P_60_PTSDF_D2': 'p_60_ptsdf_d2',
        'P_60_OZONE_D2': 'p_60_ozone_d2',
        'P_60_PM25_D2': 'p_60_pm25_d2',
    },
    'Elections': \
    {
        '1st Ballot Mailed Disposition': BALLOT_MAILED_DISPOSITION.format( 1 ),
        '1st Ballot Received By': BALLOT_RECEIVED_BY.format( 1 ),
        '1st Ballot Rejection Reason': BALLOT_REJECTION_REASON.format( 1 ),
        '1st Ballot Sent By': BALLOT_SENT_BY.format( 1 ),
        '2nd Ballot Mailed Disposition': BALLOT_MAILED_DISPOSITION.format( 2 ),
        '2nd Ballot Received By': BALLOT_RECEIVED_BY.format( 2 ),
        '2nd Ballot Rejection Reason': BALLOT_REJECTION_REASON.format( 2 ),
        '2nd Ballot Sent By': BALLOT_SENT_BY.format( 2 ),
        '3rd Ballot Mailed Disposition': BALLOT_MAILED_DISPOSITION.format( 3 ),
        '3rd Ballot Received By': BALLOT_RECEIVED_BY.format( 3 ),
        '3rd Ballot Rejection Reason': BALLOT_REJECTION_REASON.format( 3 ),
        '3rd Ballot Sent By': BALLOT_SENT_BY.format( 3 ),
        'City/ Town Code': TOWN_CODE,
        'City/ Town Code Assigned Number': TOWN_CODE,
        'City/ Town Indicator': TOWN_INDICATOR,
        'City/ Town Name': TOWN_NAME,
        'City/Town Indicator': TOWN_INDICATOR,
        'CityTown Name': TOWN_NAME,
        'Date 1st Ballot Mailed': BALLOT_MAILED_DATE.format( 1 ),
        'Date 1st Ballot Received': BALLOT_RECEIVED_DATE.format( 1 ),
        'Date 2nd Ballot Mailed': BALLOT_MAILED_DATE.format( 2 ),
        'Date 2nd Ballot Received': BALLOT_RECEIVED_DATE.format( 2 ),
        'Date 3rd Ballot Mailed': BALLOT_MAILED_DATE.format( 3 ),
        'Date 3rd Ballot Received': BALLOT_RECEIVED_DATE.format( 3 ),
        'Election Type': ELECTION_TYPE,
        'First Name': FIRST_NAME,
        'Last Name': LAST_NAME,
        'Mailing Address - Apartment Number': MADDR_APARTMENT_NUMBER,
        'Mailing Address - City/Town': MADDR_CITY,
        'Mailing Address - State': MADDR_STATE,
        'Mailing Address - Street Number/Name': MADDR_STREET,
        'Mailing Address - Zip Code': MADDR_ZIP_CODE,
        'Mailing Address Line 1': MADDR_LINE.format( 1 ),
        'Mailing Address Line 2': MADDR_LINE.format( 2 ),
        'Mailing Address Line 3': MADDR_LINE.format( 3 ),
        'Mailing Address Line 4': MADDR_LINE.format( 4 ),
        'Mailing Address Line 5': MADDR_LINE.format( 5 ),
        'Middle Name': MIDDLE_NAME,
        'Party Affiliation': PARTY_AFFILIATION,
        'Party Enrolled': PARTY_AFFILIATION,
        'Record Seq. #': PARTY_AFFILIATION,
        'Residential Address - Apartment Number': RADDR_APARTMENT_NUMBER,
        'Residential Address - Street Name': RADDR_STREET_NAME,
        'Residential Address - Street Number': RADDR_STREET_NUMBER,
        'Residential Address - Street Suffix': RADDR_STREET_NUMBER_SUFFIX,
        'Residential Address - Zip Code': RADDR_ZIP_CODE,
        'Residential Address Apt Number': RADDR_APARTMENT_NUMBER,
        'Residential Address Street Name': RADDR_STREET_NAME,
        'Residential Address Street Number': RADDR_STREET_NUMBER,
        'Residential Address Street Suffix': RADDR_STREET_NUMBER_SUFFIX,
        'Residential Address Zip Code': RADDR_ZIP_CODE,
        'Title': TITLE,
        'Type of Election': ELECTION_TYPE,
        'Vot Delete Flag': 'vote_delete_flag',
        'Voter ID': RESIDENT_ID,
        'Voter ID Number': RESIDENT_ID,
        'Voter Status': VOTER_STATUS,
        'Voter Status r': VOTER_STATUS,
        'Voter Title': TITLE,
        'AV Flag': 'av_flag',
        'Date Application Received': 'date_application_received',
        'Election Date': ELECTION_DATE,
        'EV Type': 'ev_type',
        'Party Voted': PARTY_VOTED,
        'Poll ID Required': 'poll_id_required',
        'Precinct Number': PRECINCT_NUMBER,
        'Rec Sequence Number': 'rec_sequence_number',
        'Ward Number': WARD_NUMBER,
    },
    'ElectricEesRates': \
    {
        'Year': YEAR,
        'Electric Utility': ELECTRIC_UTILITY,
        'R-1': RESIDENTIAL_R1_RATE,
        'R-2': RESIDENTIAL_R2_RATE,
        'C&I': COMMERCIAL_RATE,
        'Notes': 'notes',
        'URL': 'url',
    },
    'Employees': \
    {
        'TOWN/': DEPARTMENT,
        'LAST NAME': LAST_NAME,
        'FIRST NAME': FIRST_NAME,
        'POSITION': POSITION,
        '*REGULAR PAY': 'regular_pay',
        '*OFF-DUTY': 'off_duty_pay',
        '*OVERTIME': 'overtime_pay',
        '*OTHER': 'other_pay',
        '*RETRO': 'retroactive_pay',
        'TOTAL GROSS': 'total_gross_pay',
    },
    'GasEesRates': \
    {
        'Year': YEAR,
        'Gas Utility': GAS_UTILITY,
        'Residential': RESIDENTIAL_RATE,
        'C&I': COMMERCIAL_RATE,
        'Notes': 'notes',
        'URL': 'url',
        'Docket': 'docket',
    },
    'Gender_2014': \
    {
        'Resident Id Number': RESIDENT_ID,
        'Gender': GENDER,
    },
    'PollingPlaces': \
    {
        'Year': YEAR,
        'Precinct': 'precinct',
        'Building': 'building',
        'Address': 'address',
    },
    'RawBuildingPermits': \
    {
        'Permit#': PERMIT_NUMBER,
        'Permit Type': PERMIT_TYPE,
        'Work Description': WORK_DESCRIPTION,
        'Property Owner': OWNER_NAME,
        'Address': ADDRESS,
        'Application Date': APPLICATION_DATE,
        'Status': STATUS,
        'Applicant Name': 'applicant_name',
        'Applicant Phone#': 'applicant_phone',
        'Applicant Email': 'applicant_email',
    },
    'RawBuildingPermits_Cga': \
    {
        'City/Town': TOWN_NAME,
        'Permit #': PERMIT_NUMBER,
        'Address Num': ADDR_STREET_NUMBER,
        'Street': ADDR_STREET_NAME,
        'Contractor ID': 'contractor_id',
        'Date': DATE,
        'General Notes': 'general_notes',
        'Insp Notes': INSPECTION_NOTES,
    },
    'RawBuildingPermits_Electrical': \
    {
        'PERMIT#': PERMIT_NUMBER,
        'FILE#': FILE_NUMBER,
        'Address of the property': ADDRESS,
        'Property Owner': OWNER_NAME,
        'Work Description': WORK_DESCRIPTION,
        'Applicant': APPLICANT,
        'Status': PERMIT_STATUS,
        'Permit Fees': TOTAL_FEE,
        'Estimated Cost of Work': 'estimated_cost_$',
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
        'Last Inspection': LAST_INSPECTION_DATE,
        'Inspection Type': INSPECTION_TYPE,
        'Assigned to': ASSIGNED_TO,
        'Due (for inspection)': DATE_DUE_FOR_INSPECTION,
        'Status.1': INSPECTION_STATUS,
        'Will you be retrofitting lighting?': 'retrofit_lighting',
        'No. of Recessed Fixtures': 'recessed_fixtures',
        'No. of Lighting Outlets': 'lighting_outlets',
        'No. of Lighting Fixtures': 'lighting_fixtures',
        'No. of Receptacle Outlets': 'receptacle_outlets',
        'No. of Switches': 'switches',
        'No. of Ceiling Suspended (Paddle) Fans': 'ceiling_fans',
        'No. of Bathroom Fan/Timer': 'bathroom_fans',
        'No. of Ranges': 'ranges',
        'No. of Waste Disposers': 'waste_disposers',
        'No. of Dishwashers': 'dishwashers',
        'No. of Dryers': 'dryers',
        'No. of Water Heaters': 'water_heaters',
        'Water Heaters KW': 'water_heaters_kw',
        'No. of Oil Burners': 'oil_burners',
        'No. of Gas Burners': 'gas_burners',
        'No. of Air Conditioners': 'air_conditioners',
        'Total A/C Tons': 'total_ac_tons',
        'No. of Heat Pumps': 'heat_pumps',
        'Heat Pump Tons': 'heat_pump_tons',
        'Heat Pump KW': 'heat_pump_kw',
        'Space/Area Heating KW': 'space_heating_kw',
        'No. of Heating Appliances': 'heating_appliances',
        'Heating Appliances KW': 'heating_appliance_kw',
        'No. of Microwaves': 'microwaves',
        'Total Number of Appliances': 'total_appliances',
        'No. of Signs': 'signs',
        'No. of Ballasts': 'ballasts',
        'No. of Motors': 'motors',
        'Total Motor HP': 'total_motor_hp',
        'No. of Transformers': 'transformers',
        'Total Transformer KVA': 'total_transformer_kva',
        'No. of Generators': 'generators',
        'Total Generator KVA': 'total_generator_kva',
        'No. of Emergency Lighting Battery Units': 'lighting_batteries',
        'No. of Hydromassage Bathtubs': 'massage_tubs',
        'No. of Hot Tubs': 'hot_tubs',
        'Swimming Pool': 'swimming_pool',
        'Solar Panel Install': 'solar_panel_install',
    },
    'RawBuildingPermits_Plumbing': \
    {
        'PERMIT#': PERMIT_NUMBER,
        'FILE#': FILE_NUMBER,
        'Address of the property': ADDRESS,
        'Property Owner': OWNER_NAME,
        'Work Description': WORK_DESCRIPTION,
        'Applicant': APPLICANT,
        'Status': PERMIT_STATUS,
        'Permit Fees': TOTAL_FEE,
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
        'Last Inspection': LAST_INSPECTION_DATE,
        'Inspection Type': INSPECTION_TYPE,
        'Assigned to': ASSIGNED_TO,
        'Due (for inspection)': DATE_DUE_FOR_INSPECTION,
        'Status.1': INSPECTION_STATUS,
        'Appliance 1': APPLIANCE + '_1',
        'Appliance 2': APPLIANCE + '_2',
        'Appliance 3': APPLIANCE + '_3',
        'Appliance 4': APPLIANCE + '_4',
        'Appliance 5': APPLIANCE + '_5',
    },
    'RawBuildingPermits_Roof': \
    {
        'File#': FILE_NUMBER,
        'Permit#': PERMIT_NUMBER,
        'Applicant': APPLICANT,
        'Street Address': ADDRESS,
        'Street Zip': ZIP,
        'Descritpion': WORK_DESCRIPTION,
        'Status': PERMIT_STATUS,
        'Company Name': BUSINESS_NAME,
        'Comtact Name': CONTACT_NAME,
        'Personal Mailing Address': MAILING_ADDRESS,
        'State': MADDR_STATE,
        'Zip': MADDR_ZIP_CODE,
        'Email Address': CONTACT_EMAIL,
        'Phone #': PHONE,
        'Property Owner': OWNER_NAME,
        'Property Manager': PROPERTY_MANAGER,
        'Contractor 1': CONTRACTOR_NAME_1,
        'Contractor 2': CONTRACTOR_NAME_2,
        'Contractor 3': CONTRACTOR_NAME_3,
        'Total Project Cost': PROJECT_COST,
        'Permit Fees': TOTAL_FEE,
        'Type of Roof': ROOF_TYPE,
        'Is the entire roof being roofed?': ENTIRE_ROOF,
        'Stripping roof?': STRIPPING,
        'How many layers are currently on the roof?': ROOF_LAYERS,
        'Submittal Documents': SUBMITTAL_DOCUMENTS,
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
    },
    'RawBuildingPermits_Siding': \
    {
        'File#': FILE_NUMBER,
        'Permit#': PERMIT_NUMBER,
        'Applicant': APPLICANT,
        'Street Address': ADDRESS,
        'Street Zip': ZIP,
        'Descritpion': WORK_DESCRIPTION,
        'Status': PERMIT_STATUS,
        'Company Name': BUSINESS_NAME,
        'Comtact Name': CONTACT_NAME,
        'Personal Mailing Address': MAILING_ADDRESS,
        'State': MADDR_STATE,
        'Zip': MADDR_ZIP_CODE,
        'Email Address': CONTACT_EMAIL,
        'Phone #': PHONE,
        'Property Owner': OWNER_NAME,
        'Property Manager': PROPERTY_MANAGER,
        'Contractor 1': CONTRACTOR_NAME_1,
        'Contractor 2': CONTRACTOR_NAME_2,
        'Contractor 3': CONTRACTOR_NAME_3,
        'Total Project Cost': PROJECT_COST,
        'Permit Fees': TOTAL_FEE,
        'Type of Siding': SIDING_TYPE,
        'Submittal Documents': SUBMITTAL_DOCUMENTS,
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
    },
    'RawBuildingPermits_Solar': \
    {
        'PERMIT#': PERMIT_NUMBER,
        'FILE#': FILE_NUMBER,
        'Address of the property': ADDRESS,
        'Property Owner': OWNER_NAME,
        'Work Description': WORK_DESCRIPTION,
        'Applicant': APPLICANT,
        'Status': PERMIT_STATUS,
        'Permit Fees': TOTAL_FEE,
        'Use Group': PROPERTY_USE_GROUP,
        'Total project cost': PROJECT_COST,
        'Number of solar PV modules': MODULES,
        'Watts per module': WATTS_PER_MODULE,
        'Number of interactive inverters': INTERACTIVE_INVERTERS,
        'Power ratings (each)': POWER_RATINGS_EACH,
        'Electrical permit number': ELECTRICAL_PERMIT_NUMBER,
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
        'Last Inspection': LAST_INSPECTION_DATE,
        'Inspection Type': INSPECTION_TYPE,
        'Assigned to': ASSIGNED_TO,
        'Due (for inspection)': DATE_DUE_FOR_INSPECTION,
        'Status.1': INSPECTION_STATUS,
    },
    'RawBuildingPermits_Sunrun': \
    {
        'STATUS': STATUS,
        'STATUS_1': STATUS + '_1',
        'ADDRESS': ADDRESS,
        'ADDRESS_1': ADDRESS + '_1',
        'OPENED': OPENED,
        'CLOSED': CLOSED_DATE,
    },
    'RawBuildingPermits_Wx': \
    {
        'File#': FILE_NUMBER,
        'Address': ADDRESS,
        'Project Description': WORK_DESCRIPTION,
        'Permit#': PERMIT_NUMBER,
        'Permit Type': PERMIT_TYPE,
        'Subtype': PERMIT_SUBTYPE,
        'Applicant': APPLICANT,
        'Business Name': BUSINESS_NAME,
        'Mailing Address': MADDR_STREET,
        'Mailing CSZ': MADDR_CSZ,
        'Permit Status': PERMIT_STATUS,
        'Permit Fees': TOTAL_FEE,
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
        'Property Owner': OWNER_NAME,
        'Use Group': PROPERTY_USE_GROUP,
        'Total Project Cost': PROJECT_COST,
    },
    'RawBuildingPermits_Wx_Ongoing': \
    {
        'Permit#': PERMIT_NUMBER,
        'File#': FILE_NUMBER,
        'Address': ADDRESS,
        'Zip': ZIP,
        'Permit Type': PERMIT_TYPE,
        'Subtype': PERMIT_SUBTYPE,
        'Work Description': WORK_DESCRIPTION,
        'Applicant': APPLICANT,
        'Business Name': BUSINESS_NAME,
        'Mailing Address': MADDR_STREET,
        'Mailing CSZ': MADDR_CSZ,
        'Permit Status': PERMIT_STATUS,
        'Permit Fees': TOTAL_FEE,
        'Application Date': APPLICATION_DATE,
        'Approval Date': APPROVAL_DATE,
        'Issue Date': DATE_ISSUED,
        'Expiration Date': EXPIRATION_DATE,
        'Close Date': CLOSED_DATE,
        'Property Owner': OWNER_NAME,
        'Use Group': PROPERTY_USE_GROUP,
        'Total Project Cost': PROJECT_COST,
    },
    'RawBuildingPermits_Wx_Past': \
    {
        'Info_1': FILE_NUMBER,
        'Permit#': PERMIT_NUMBER,
        'Status': PERMIT_STATUS,
        'Address': ADDRESS,
        'Opened': DATE_ISSUED,
        'Closed': CLOSED_DATE,
        'Permit#2': PERMIT_NUMBER + '_2',
        'Permit#3': PERMIT_NUMBER + '_3',
        'Description': WORK_DESCRIPTION,
    },
    'RawBusinesses_1': \
    {
        'License#': LICENSE_NUMBER,
        'Business Name': BUSINESS_NAME,
        'Number of Employees': EMPLOYEES,
        'License Type': 'license_type',
        'Subtype': LICENSE_SUBTYPE,
        'Property Address': LOCATION,
        'Applicant': APPLICANT,
        'Email': EMAIL,
        'Phone Number': PHONE,
        'Address': APPLICANT_ADDRESS,
        'Application Date': APPLICATION_DATE,
        'Closed Date': CLOSED_DATE,
        'Status': STATUS,
    },
    'RawBusinesses_2': \
    {
        'License#': LICENSE_NUMBER,
        'Applicant': APPLICANT,
        'Property Address': LOCATION,
        'Status': STATUS,
        'Applicant Address': APPLICANT_ADDRESS,
        'Phone Number(s)': PHONE,
        'Email Address': EMAIL,
        'Num Employees': EMPLOYEES,
        'Business Manager': BUSINESS_MANAGER,
        'SBR#': 'sbr_number',
    },
    'RawCensus_L': \
    {
        'Census Yr': CENSUS_YEAR,
        'Res. ID': RESIDENT_ID,
        'Last Nm': LAST_NAME,
        'First Nm': FIRST_NAME,
        'Middle Nm': MIDDLE_NAME,
        'Title (Jr, Sr, II, III, etc.)': TITLE,
        'D.O.B. (mm/dd/yyyy format)': DATE_OF_BIRTH,
        'Occupation': OCCUPATION,
        'Mail to Code (Y if HOH; N if Not)': HEAD_OF_HOUSEHOLD,
        'Family Group Code (Within an address)': FAMILY_GROUP_CODE,
        'Res. - St. #': RADDR_STREET_NUMBER,
        'Res. - St. # Suffix': RADDR_STREET_NUMBER_SUFFIX,
        'Res. - St. Name': RADDR_STREET_NAME,
        'Res. - Apt #': RADDR_APARTMENT_NUMBER,
        'Res. - Zip Code': RADDR_ZIP_CODE,
        'Mail - Street # and Name': MADDR_STREET,
        'Mail - Apt #': MADDR_APARTMENT_NUMBER,
        'Mail - City and Town': MADDR_CITY,
        'Mail - State': MADDR_STATE,
        'Mail - Zip Code': MADDR_ZIP_CODE,
        'Phone #': PHONE,
        'Party': PARTY_AFFILIATION,
        'Ward #': WARD_NUMBER,
        'Precinct #': PRECINCT_NUMBER,
        'Voter Status (A- Active, I - Inactive (if voter))': VOTER_STATUS,
        'Vet. flag (Y if vet)': VETERAN,
    },
    'RawCommercial_1': \
    {
        'Account Number': ACCOUNT_NUMBER,
        'Gis ID': GIS_ID,
        'Location': LOCATION,
        'Owners Name': OWNER_NAME,
        'Address Line 1': MADDR_LINE.format( 1 ),
        'City': MADDR_CITY,
        'State': MADDR_STATE,
        'Zip': MADDR_ZIP_CODE,
        'Deed Book': BOOK,
        'Deed Page': PAGE,
        'Sale Date': SALE_DATE,
        'Style:': STYLE,
        'Style Desc': STYLE + _DESC,
        'Stories:': STORY_HEIGHT,
        'Stories Desc': STORY_HEIGHT + _DESC,
        'Occupancy': RENTAL_LIVING_UNITS,
        'Roof Structure': ROOF_STRUCTURE,
        'Usrfld 201 Desc': ROOF_STRUCTURE + _DESC,
        'Roof Cover': ROOF_COVER,
        'Usrfld 202 Desc': ROOF_COVER + _DESC,
        'Heating Fuel': HEATING_FUEL,
        'Usrfld 207 Desc': HEATING_FUEL + _DESC,
        'Heating Type': HEATING_TYPE,
        'Usrfld 208 Desc': HEATING_TYPE + _DESC,
        'AC Type': AC_TYPE,
        'Usrfld 209 Desc': AC_TYPE + _DESC,
        'Total Bedrms': BEDROOOMS,
        'Usrfld 216 Desc': BEDROOOMS + _DESC,
        'Total Baths': BATHS,
        'Usrfld 217 Desc': BATHS + _DESC,
        'Bldg Cond': BUILDING_CONDITION,
        'Usrfld 219 Desc': BUILDING_CONDITION + _DESC,
        'Total Assessed Parcel Value': TOTAL_ASSESSED_VALUE,
        'Prop Class': 'property_class',
    },
    'RawCommercial_2': \
    {
        'REM_ACCT_NUM': ACCOUNT_NUMBER,
        'REM_USE_CODE': LAND_USE_CODE,
        'CNS_OCC': LAND_USE_CODE + '_1',
        'CNS_OCC_DESC': LAND_USE_CODE + _DESC,
    },
    'RawContracts': \
    {
        'Contract': CONTRACT,
        'Method': METHOD,
        'Status': STATUS,
        'Description': DESCRIPTION,
        'Year': YEAR,
        'Period': PERIOD,
        'Revised Amount': AMOUNT,
        'Vendor': VENDOR,
        'Dept': DEPARTMENT,
    },
    'RawElectricUsage': \
    {
        'year': YEAR,
        'Town': TOWN_NAME,
        'Sector': SECTOR,
        'Jan': JAN,
        'Feb': FEB,
        'Mar': MAR,
        'Apr': APR,
        'May': MAY,
        'Jun': JUN,
        'Jul': JUL,
        'Aug': AUG,
        'Sept': SEP,
        'Oct': OCT,
        'Nov': NOV,
        'Dec': DEC,
        'Annual': ANNUAL_ELECTRIC_USAGE,
    },
    'RawExternalSuppliersElectric_L': \
    {
        'CustNo_i': 'customer_number',
        'CustomerName': CUSTOMER_NAME,
        'LDCAcctNo_vc': ACCOUNT_NUMBER,
        'Acctno_i': ACCOUNT_NUMBER + '_i',
        'LDC_vc': 'ldc',
        'AcctStatus': 'account_status',
        'Address': LOCATION_ADDRESS,
        'TurnOnDate': 'turn_on_date',
        'TurnOffDate': 'turn_off_date',
        'OfferContract_vc': 'offer_contract',
        'ContractStatus': 'contract_status',
        'ContractID': 'contract_id',
        'PlanName': 'plan_name',
        'OPP_WonBid_ID': 'opp_won_bid_id',
        'OriginationDate': 'origination_date',
        'Effective_Start': 'effective_start_date',
        'Effective_End': 'effective_end_date',
        'Origination_Start': 'origination_start_date',
        'Origination_End': 'origination_end_date',
        'PMT_Term': 'pmt_term',
        'InvoiceResponsible_vc': 'invoice_responsible',
    },
    'RawExternalSuppliersGas_L': \
    {
        'Customer Name': CUSTOMER_NAME,
        'Commodity': 'commodity',
        'Billing Account': 'billing_account',
        'Utility Account': UTILITY_ACCOUNT,
        'Address': STREET,
        'City': CITY,
        'State': STATE,
        'Zip': ZIP,
        'Status': STATUS,
        'Service Start Date': 'service_start_date',
        'Service End Date': 'service_end_date',
    },
    'RawGasUsage': \
    {
        'year': YEAR,
        'Town': TOWN_NAME,
        'Sector': SECTOR,
        'Jan': JAN,
        'Feb': FEB,
        'Mar': MAR,
        'Apr': APR,
        'May': MAY,
        'Jun': JUN,
        'Jul': JUL,
        'Aug': AUG,
        'Sept': SEP,
        'Oct': OCT,
        'Nov': NOV,
        'Dec': DEC,
        'Annual': ANNUAL_GAS_USAGE,
    },
    'RawGeographicReport': \
    {
        'year': YEAR,
        'County': COUNTY,
        'Town': TOWN_NAME,
        'Zip Code': ZIP,
        'Sector': SECTOR,
        'Annual Electric Usage (MWh)': ANNUAL_ELECTRIC_USAGE,
        'Annual Electric Savings (MWh)': ANNUAL_ELECTRIC_SAVINGS,
        'Electric Incentives': ELECTRIC_INCENTIVES,
        'Annual Gas Usage (Therms)': ANNUAL_GAS_USAGE,
        'Annual Gas Savings (Therms)': ANNUAL_GAS_SAVINGS,
        'Gas Incentives': GAS_INCENTIVES,
    },
    'RawGlcacJobs': \
    {
        'Job Num': JOB_NUMBER,
        'Address': ADDRESS,
        'Completed Date': 'completed_date',
        'Residence Type': 'residence_type',
        'Ownership': 'ownership',
        'Fuel Type': HEATING_FUEL,
        'DHW Fuel Type': 'dhw_fuel',
        'First Name': FIRST_NAME,
        'Last Name': LAST_NAME,
        'Phone Number': PHONE,
        'Utility Compensation': 'utility_compensation_$',
        'Total Job Cost': PROJECT_COST,
        'Diffrence': 'cost_difference_$',
        'Residence Status': STATUS,
        'Uitlity': 'utility',
    },
    'RawIsoZones_L': \
    {
        'account_number': ACCOUNT_NUMBER,
        'iso_zone': ISO_ZONE,
    },
    'RawLocalElectionResults': \
    {
        'P-1': PRECINCT_1,
        'P-2': PRECINCT_2,
        'P-3': PRECINCT_3,
        'P-4': PRECINCT_4,
        'P-5': PRECINCT_5,
        'P-6': PRECINCT_6,
        'P-7/7A': PRECINCT_7,
        'P-8': PRECINCT_8,
        'P-9': PRECINCT_9,
        'election_date': ELECTION_DATE,
        'PRECINCTS:': OFFICE_OR_CANDIDATE,
    },
    'RawMassEnergyInsight_A': \
    {
        'Department': DEPARTMENT,
        'Complex': 'complex',
        'Facility': 'facility',
        'Fuel (units)': 'fuel_units',
        'Account #': ACCOUNT_NUMBER,
        'Usage End Date Null': USAGE_END_NULL,
        'One': ONE,
        'Provider': 'provider',
        'Unnamed: 8': COST_OR_USE,
    },
    'RawResidential_1': \
    {
        'REM_ACCT_NUM': ACCOUNT_NUMBER,
        'REM_GIS_ID': GIS_ID,
        'REM_OWN_NAME': OWNER_NAME,
        'REM_PRCL_LOCN': LOCATION,
        'PRC_TTL_ASSESS': TOTAL_ASSESSED_VALUE,
        'REM_USE_CODE': LAND_USE_CODE,
        'OWN_NAME1': OWNER_1_NAME,
        'OWN_NAME2': OWNER_2_NAME,
        'MAD_MAIL_ADDR1': MADDR_LINE.format( 1 ),
        'MAD_MAIL_ADDR2': MADDR_LINE.format( 2 ),
        'MAD_MAIL_CITY': MADDR_CITY,
        'MAD_MAIL_STATE': MADDR_STATE,
        'MAD_MAIL_ZIP': MADDR_ZIP_CODE,
        'SLH_SALE_DATE': SALE_DATE,
        'SLH_BOOK': BOOK,
        'SLH_PAGE': PAGE,
        'SLH_PRICE': SALE_PRICE,
        'CNS_STORIES': STORY_HEIGHT,
        'CNS_STORIES_DESC': STORY_HEIGHT + _DESC,
        'CNS_ROOF_COVER': ROOF_COVER,
        'CNS_ROOF_COVER_DESC': ROOF_COVER + _DESC,
        'CNS_ROOF_STRUCT': ROOF_STRUCTURE,
        'CNS_ROOF_STRUCT_DESC': ROOF_STRUCTURE + _DESC,
        'CNS_USRFLD_82': HEATING_FUEL + _CODE,
        'CNS_USRFLD_82_DESC': HEATING_FUEL,
        'CNS_HEAT_TYPE': HEATING_TYPE,
        'CNS_HEAT_TYPE_DESC': HEATING_TYPE + _DESC,
        'CNS_AC_TYPE': AC_TYPE,
        'CNS_AC_TYPE_DESC': AC_TYPE + _DESC,
        'CNS_NUM_BEDRM': BEDROOOMS,
        'CNS_NUM_BEDRM_DESC': BEDROOOMS + _DESC,
        'CNS_NUM_BATHRM': FULL_BATHS,
        'CNS_NUM_BATHRM_DESC': FULL_BATHS + _DESC,
        'CNS_NUM_HALF_BATHS': HALF_BATHS,
        'CNS_NUM_RMS': ROOMS,
        'CNS_NUM_RMS_DESC': ROOMS + _DESC,
        'CNS_KITCHEN_STYLE': KITCHEN_STYLE,
        'CNS_KITCHEN_STYLE_DESC': KITCHEN_STYLE + _DESC,
    },
    'RawResidential_2': \
    {
        'REM_ACCT_NUM': ACCOUNT_NUMBER,
        'REM_GIS_ID': GIS_ID,
        'REM_OWN_NAME': OWNER_NAME,
        'REM_PRCL_LOCN': LOCATION,
        'CNS_AYB': YEAR_BUILT,
        'CNS_EYB': EFFECTIVE_YEAR_BUILT,
        'CNS_STATUS': STATUS,
        'CNS_YEAR_REMODEL': YEAR_REMODELED,
        'CNS_BASEMENT': BASEMENT,
        'CNS_BASEMENT_DESC': BASEMENT + _DESC,
        'CNS_BATHRM_STYLE': BATHROOM_STYLE,
        'CNS_BATHRM_STYLE_DESC': BATHROOM_STYLE + _DESC,
        'CNS_BATHRM_STYLE2': BATHROOM_STYLE + '_2',
        'CNS_BATHRM_STYLE2_DESC': BATHROOM_STYLE + '_2' + _DESC,
        'CNS_FIREPLACES': FIREPLACES,
        'CNS_HEAT_TYPE': HEATING_TYPE,
        'CNS_HEAT_TYPE_DESC': HEATING_TYPE + _DESC,
        'CNS_USRFLD_82': HEATING_FUEL + _CODE,
        'CNS_USRFLD_82_DESC': HEATING_FUEL,
        'CNS_KITCHEN_STYLE': KITCHEN_STYLE,
        'CNS_KITCHEN_STYLE_DESC': KITCHEN_STYLE + _DESC,
        'CNS_KITCHEN_STYLE2': KITCHEN_STYLE + '_2',
        'CNS_KITCHEN_STYLE2_DESC': KITCHEN_STYLE + '_2' + _DESC,
        'CNS_KITCHEN_STYLE3': KITCHEN_STYLE + '_3',
        'CNS_NUM_BATHRM': FULL_BATHS,
        'CNS_NUM_BEDRM': BEDROOOMS,
        'CNS_NUM_HALF_BATHS': HALF_BATHS,
        'CNS_ROOF_COVER': ROOF_COVER,
        'CNS_ROOF_COVER_DESC': ROOF_COVER + _DESC,
        'CNS_ROOF_STRUCT': ROOF_STRUCTURE,
        'CNS_ROOF_STRUCT_DESC': ROOF_STRUCTURE + _DESC,
    },
    'RawResidential_3': \
    {
        'Account Number': ACCOUNT_NUMBER,
        'Gis ID': GIS_ID,
        'Owners Name': OWNER_NAME,
        'Location': LOCATION,
        'Year Built': YEAR_BUILT,
        'Effective Year Built': EFFECTIVE_YEAR_BUILT,
        'Condition': BUILDING_CONDITION,
        'Year Remodeled': YEAR_REMODELED,
        'Basement': BASEMENT,
        'Bath Style:': BATHROOM_STYLE,
        'Fireplaces': FIREPLACES,
        'Heat Type:': HEATING_TYPE,
        'Kitchen Style:': KITCHEN_STYLE,
        'Kitchen Style Desc': KITCHEN_STYLE + _DESC,
        'Total Bthrms:': FULL_BATHS,
        'Total Bedrooms:': BEDROOOMS,
        'Total Half Baths:': HALF_BATHS,
        'Total Rooms:': ROOMS,
        'Roof Cover': ROOF_COVER,
        'Roof Cover Desc': ROOF_COVER + _DESC,
        'Roof Structure:': ROOF_STRUCTURE,
        'Roof Struct Desc': ROOF_STRUCTURE + _DESC,
        'Style:': STYLE,
    },
    'RawResidential_4': \
    {
        'ACCOUNT NUM': ACCOUNT_NUMBER,
        'GIS ID': GIS_ID,
        'OWNER': OWNER_NAME,
        'STREET NUM': LADDR_STREET_NUMBER,
        'STREET NAME': LADDR_STREET_NAME,
        'ASSESS CODE 1': LAND_USE_CODE,
        'GRANTEE 1': OWNER_1_NAME,
        'GRANTEE 2': OWNER_2_NAME,
        'ADDRESS1': MADDR_LINE.format( 1 ),
        'ADDRESS2': MADDR_LINE.format( 2 ),
        'CITY': MADDR_CITY,
        'STATE': MADDR_STATE,
        'ZIP': MADDR_ZIP_CODE,
        'SALE DATE 1': SALE_DATE,
        'BOOK PAGE 1': BOOK,
        'SALE PRICE 1': SALE_PRICE,
        'Stories:': STORY_HEIGHT,
        'Stories Desc': STORY_HEIGHT + _DESC,
        'Roof Cover': ROOF_COVER,
        'Roof Cover Desc': ROOF_COVER + _DESC,
        'Roof Structure:': ROOF_STRUCTURE,
        'Roof Struct Desc': ROOF_STRUCTURE + _DESC,
        'Heat Type:': HEATING_FUEL + _CODE,
        'Usrfld 82 Desc': HEATING_FUEL,
        'Heat Fuel': HEATING_TYPE,
        'Heat Type Desc': HEATING_TYPE + _DESC,
        'AC Type:': AC_TYPE,
        'Ac Type Desc': AC_TYPE + _CODE,
        'Total Bedrooms:': BEDROOOMS,
        'Num Bedrm Desc': BEDROOOMS + _DESC,
        'Total Bthrms:': FULL_BATHS,
        'Num Bathrm Desc': FULL_BATHS + _DESC,
        'Total Half Baths:': HALF_BATHS,
        'Total Rooms:': ROOMS,
        'Num Rms Desc': ROOMS + _DESC,
        'Kitchen Style:': KITCHEN_STYLE,
        'Kitchen Style Desc': KITCHEN_STYLE + _DESC,
    },
    'RawResidential_5': \
    {
        'ACCOUNT NUM': ACCOUNT_NUMBER,
        'VISION ID': VISION_ID,
        'TOTAL LAND AREA': TOTAL_ACRES,
        'PARCEL TTL ASSESSED': 'title_assessed_$',
        'PARCEL BLDG ASSESSED': 'building_assessed_$',
        'PARCEL LAND APPRAISED': 'land_assessed_$',
        'PARCEL OUT BLDG ASSESSED': 'outbuilding_assessed_$',
        'PARCEL XTRA FEATURE ASSESSED': 'extra_feature_assessed_$',
        'PARCEL BLDG COUNT': BUILDING_COUNT,
        'PARCEL OUTBUILDING COUNT': 'outbuilding_count',
        'STYLE': STYLE,
        'GROSS BULDING AREA': GROSS_BUILDING_AREA,
        'LIVING AREA': LIVING_AREA,
        'EFFECTIVE AREA': 'effective_area',
        'Num Kitchens': KITCHENS,
        'Occupancy': OCCUPANCY_HOUSEHOLDS,
        'Exterior Wall 1': EXTERIOR_WALL_TYPE,
        'Ext Wall 1 Desc': EXTERIOR_WALL_TYPE + _DESC,
        'Grade:': GRADE,
        'Grade Desc': GRADE + _DESC,
        'Interior Flr 1': 'interior_floor_1',
        'View Desc': 'interior_floor_1' + _DESC,
        'Interior Flr 2': 'interior_floor_2',
        'Bldg Cond': BUILDING_CONDITION,
        'Usrfld 101 Desc': BUILDING_CONDITION + _DESC,
    },
    'RawVendors': \
    {
        'VDR NUM': VENDOR_NUMBER,
        'NAME': VENDOR_NAME,
        'DBA': 'dba',
        'ADDR1': ADDR_1,
        'ADDR2': ADDR_2,
        'CITY': CITY,
        'ST': STATE,
        'ZIP': ZIP,
        'ALPHA': 'alpha',
        'EMAIL': EMAIL,
        'CONTACT1 NAME': CONTACT_NAME,
        'CONTACT1 EMAIL': CONTACT_EMAIL,
        'CONTACT1 PHONE': PHONE,
    },
    'Solar': \
    {
        'FIRST_NAME': FIRST_NAME,
        'MIDDLE_NAME': MIDDLE_NAME,
        'LAST_NAME': LAST_NAME,
        'SITE_ADDR': SITE_ADDRESS,
        'TOWN': 'town',
        'F1_EST_KW': 'estimated_kw',
        'CREDIT': 'credit',
        'ELEC_UTIL': 'electric_utility',
        'EFFCTV_RATE': 'effective_rate',
        'EST_KWH_CN': 'estimated_kwh',
        'MAX_BEHAV_SCORE': 'max_behavior_score',
        'PHONE': PHONE,
    },
    'Water': \
    {
        'serv_id': SERVICE_ID,
        'acct_no': ACCOUNT_NUMBER,
        'service': SERVICE,
        'last': LAST_NAME,
        'first': FIRST_NAME,
        'number': ADDR_STREET_NUMBER,
        'name': ADDR_STREET_NAME,
        'meter_no': METER_NUMBER,
        'desc': DESCRIPTION,
        'tran_date': TRANSACTION_DATE,
        'type_str': TRANSACTION_TYPE,
        'units': UNITS,
        'prior_date': PRIOR_DATE,
        'cur_date': CURRENT_DATE,
        'cur_read': CURRENT_READING,
        'prior_read': PRIOR_READING,
        'service_type': SERVICE_TYPE,
    },
}
CONSISTENT_COLUMN_NAMES['RawBuildingPermits_Gas'] = CONSISTENT_COLUMN_NAMES['RawBuildingPermits_Plumbing']
CONSISTENT_COLUMN_NAMES['RawMassEnergyInsight_L'] = CONSISTENT_COLUMN_NAMES['RawMassEnergyInsight_A']


# Calculate age of a person
def calculate_age( date_of_birth ):
    age = datetime.date.today().year - int( date_of_birth.split( '-' )[0] ) if date_of_birth else 0
    return int( age )


# Clean  residential land use codes for reliable matching
def clean_residential_land_use_codes( sr_res_codes ):
    sr_res_codes = sr_res_codes.astype(str).str.zfill( 4 )
    return sr_res_codes


# Generate list of residential land use codes, used to categorize properties
def get_residential_land_use_codes( luc_filename ):
    df_res_codes = pd.read_excel( luc_filename, dtype=object )
    sr_res_codes = df_res_codes['Residential Land Use Code']
    sr_res_codes = clean_residential_land_use_codes( sr_res_codes )
    sr_res_codes = sr_res_codes.drop_duplicates()
    ls_res_codes = list( sr_res_codes )
    return ls_res_codes


# --> Generate column name mappings for Mass Energy Insight tables -->
def populate_mei_column_names( dict, start_year, end_year ):

    fy_month_base = 6 # End of fiscal year
    months_per_year = 12
    start_month = months_per_year
    end_month = start_month + months_per_year

    for n_year in range( start_year, end_year ):

        for n in range( start_month, end_month ):

            # Format month number
            n_month = 1 + ( n % 12 )
            month_number = str( n_month )

            # Derive month name from month number
            datetime_object = datetime.datetime.strptime( month_number, '%m' )
            month_name = datetime_object.strftime( '%B' )

            # Optionally add entry to dictionary
            if ( n_month > fy_month_base ) or ( n_year > start_year ):

                # Generate suffix to match column names produced by pandas
                suffix_number = n_year - start_year
                if n_month <= fy_month_base:
                    suffix_number = suffix_number - 1
                if suffix_number > 0:
                    month_name = month_name + '.'  + str( suffix_number )

                dict[month_name] = str( n_year ) + '_' + month_number.zfill( 2 )

######
# Hard-coded calls, no longer needed because mass_energy_insight.py generates calls dynamically:
# populate_mei_column_names( CONSISTENT_COLUMN_NAMES['RawMassEnergyInsight_A'], 2011, 2030 )
# populate_mei_column_names( CONSISTENT_COLUMN_NAMES['RawMassEnergyInsight_L'], 2011, 2030 )
######

# <-- Generate column name mappings for Mass Energy Insight tables <--



COLUMN_GROUP = \
{
    'PERMIT_IDS':
    [
        PERMIT_NUMBER,
        FILE_NUMBER,
    ],
    'PERMIT_DATES':
    [
        APPLICATION_DATE,
        APPROVAL_DATE,
        DATE_ISSUED,
        DATE_DUE_FOR_INSPECTION,
        LAST_INSPECTION_DATE,
        CLOSED_DATE,
        EXPIRATION_DATE,
    ],
    'PERMIT_DATES_1':
    [
        APPLICATION_DATE,
        APPROVAL_DATE,
        DATE_ISSUED,
        CLOSED_DATE,
        EXPIRATION_DATE,
    ],
    'RESIDENT':
    [
        RESIDENT_ID,
        AGE,
        GENDER,
        LAST_NAME,
        FIRST_NAME,
        OCCUPATION,
        VOTED,
        PARTY_AFFILIATION,
    ],
    'VOTER':
    [
        VOTER_ENGAGEMENT_SCORE,
        PRIMARY_ELECTIONS_VOTED,
        LOCAL_ELECTIONS_VOTED,
        STATE_ELECTIONS_VOTED,
        SPECIAL_ELECTIONS_VOTED,
        TOWN_MEETINGS_ATTENDED,
        EARLY_VOTES,
        PRECINCT_NUMBER,
    ],
    'SCORE':
    [
        PARTY_PREFERENCE_SCORE,
        LEGACY_PREFERENCE_SCORE,
        LIKELY_DEM_SCORE,
        LEGACY_DEM_SCORE,
    ],
    'ADDRESS':
    [
        NORMALIZED_STREET_NUMBER,
        RADDR_STREET_NUMBER_SUFFIX,
        LADDR_ALT_STREET_NUMBER,
        NORMALIZED_STREET_NAME,
        RADDR_APARTMENT_NUMBER,
    ],
    'HOME':
    [
        TOTAL_ASSESSED_VALUE,
        IS_HOMEOWNER,
        IS_FAMILY,
    ],
    'NORMALIZED_ADDRESS_PARTS':
    [
        NORMALIZED_ADDRESS,
        NORMALIZED_STREET_NUMBER,
        NORMALIZED_STREET_NAME,
        NORMALIZED_OCCUPANCY,
        NORMALIZED_ADDITIONAL_INFO,
    ],
}

COLUMN_ORDER = \
{
    'Assessment':
    [
        PARCEL_ID,
    ],
    'Assessment_L_Commercial':
    [
        'vision_id',
        'account_number',
        'mblu',
        'primary_land_use_code',
        'primary_land_use_code_description',
        'first_floor_use',
        'zone',
        'age',
        'street_name',
        'street_number',
        'location',
        'additional_address_info',
        'building_count',
        'total_assessed_value',
        'building_assessed_$',
        'sale_price',
        'sale_date',
        'title_assessed_$',
        'owner_1_name',
        'owner_2_name',
        'mailing_address_line_1',
        'mailing_address_line_2',
        'mailing_address_city',
        'mailing_address_state',
        'mailing_address_zip_code',
        'story_height',
        'story_height_description',
        'year_built',
        'effective_year_built',
        'total_acres',
        'land_assessed_$',
        'status',
        'style',
        'roof_cover',
        'roof_cover_description',
        'roof_structure',
        'roof_structure_description',
        'heating_fuel_code',
        'heating_fuel',
        'heating_type',
        'heating_type_description',
        'heating_fuel_description',
        'ac_type',
        'ac_type_description',
        'heat_ac',
        'number_of_kitchens',
        'number_of_bedrooms',
        'occupancy_households',
        'number_of_baths',
        'residential_units',
        'total_baths',
        'total_kitchens',
        'total_occupancy',
        'number_of_rooms',
        'number_of_rooms_description',
        'number_of_full_baths',
        'number_of_half_baths',
        'bathroom_style',
        'bathroom_style_description',
        'number_of_bedrooms_description',
        'kitchen_style',
        'kitchen_style_description',
        'living_area',
        'effective_area',
        'building_condition',
        'building_condition_description',
        'interior_floor_1',
        'interior_floor_1_description',
        'interior_floor_2',
        'exterior_wall_type',
        'exterior_wall_type_description',
        'extra_feature_assessed_$',
        'grade',
        'grade_description',
        'gross_building_area',
        'occupancy',
        'owner_name',
        'gis_id',
        'outbuilding_assessed_$',
        'outbuilding_count',
        'property_class',
        'book',
        'page',
        'rental_living_units',
        'style_description',
        'normalized_address',
        'primary_land_use_code_1',
    ],
    'Assessment_L_Residential':
    [
        # Same as Assessment_L_Commercial
    ],
    'BuildingPermits_L':
    [
        PERMIT_NUMBER,
        APPLICATION_DATE,
        * COLUMN_GROUP['NORMALIZED_ADDRESS_PARTS'],
        STATUS,
    ],
    'BuildingPermits_L_Cga':
    [
        PERMIT_NUMBER,
        DATE,
        * COLUMN_GROUP['NORMALIZED_ADDRESS_PARTS'],
        TOWN_NAME,
    ],
    'BuildingPermits_L_Electrical':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
    ],
    'BuildingPermits_L_Gas':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
    ],
    'BuildingPermits_L_Plumbing':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
    ],
    'BuildingPermits_L_Roof':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
    ],
    'BuildingPermits_L_Siding':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
    ],
    'BuildingPermits_L_Solar':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
        ADDRESS,
        APPLICANT,
        PERMIT_STATUS,
        KW_DC,
        MODULES,
        WATTS_PER_MODULE,
        INTERACTIVE_INVERTERS,
        ELECTRICAL_PERMIT_NUMBER,
        TOTAL_FEE,
        PROJECT_COST,
        PROPERTY_USE_GROUP,
        POWER_RATINGS_EACH,
        OWNER_NAME,
        * COLUMN_GROUP['NORMALIZED_ADDRESS_PARTS'],
        INSPECTION_TYPE,
        ASSIGNED_TO,
        INSPECTION_STATUS,
        WORK_DESCRIPTION,
        ACCOUNT_NUMBER,
    ],
    'BuildingPermits_L_Sunrun':
    [
        PERMIT_NUMBER,
        PERMIT_NUMBER_LINK,
        FILE_NUMBER,
        FILE_NUMBER_LINK,
        CLOSED_DATE,
        * COLUMN_GROUP['NORMALIZED_ADDRESS_PARTS'],
    ],
    'BuildingPermits_L_Wx':
    [
        * COLUMN_GROUP['PERMIT_IDS'],
        NORMALIZED_STREET_NAME,
        NORMALIZED_STREET_NUMBER,
        PERMIT_TYPE,
        PERMIT_SUBTYPE,
        WORK_DESCRIPTION,
    ],
    'Businesses_L':
    [
        LICENSE_NUMBER,
        ACCOUNT_NUMBER,
        MBLU,
        BUSINESS_NAME,
        LICENSE_SUBTYPE,
        STATUS,
        LOCATION,
        APPLICANT,
        EMAIL,
        PHONE,
        APPLICANT_ADDRESS,
        APPLICATION_DATE,
        * COLUMN_GROUP['NORMALIZED_ADDRESS_PARTS'],
        # RENTAL_LIVING_UNITS,
        # EMPLOYEES,
        # OWNER_NAME,
        # BUSINESS_MANAGER,
        # MADDR_LINE.format( 1 ),
        # MADDR_CITY,
        # MADDR_STATE,
        # MADDR_ZIP_CODE,
        # HEATING_FUEL,
        # HEATING_FUEL + _DESC,
        # HEATING_TYPE,
        # HEATING_TYPE + _DESC,
        # AC_TYPE,
        # AC_TYPE + _DESC,
        # TOTAL_ASSESSED_VALUE,
        # LAND_USE_CODE,
        # LAND_USE_CODE + '_1',
        # LAND_USE_CODE + _DESC,
        # STORY_HEIGHT,
        # ROOF_STRUCTURE,
        # ROOF_STRUCTURE + _DESC,
        # ROOF_COVER,
        # ROOF_COVER + _DESC,
        # SALE_DATE,
        # LICENSE_TYPE,
        # CLOSED_DATE,
    ],
    'Census':
    [
        RESIDENT_ID,
    ],
    'Census_L':
    [
        RESIDENT_ID,
        LAST_NAME,
        FIRST_NAME,
        MIDDLE_NAME,
        TITLE,
        * COLUMN_GROUP['NORMALIZED_ADDRESS_PARTS'],
        DATE_OF_BIRTH,
        AGE,
        OCCUPATION,
        VETERAN,
        HEAD_OF_HOUSEHOLD,
        FAMILY_GROUP_CODE,
        RADDR_STREET_NUMBER,
        RADDR_STREET_NUMBER_SUFFIX,
        RADDR_STREET_NAME,
        RADDR_APARTMENT_NUMBER,
        RADDR_ZIP_CODE,
        MADDR_STREET,
        MADDR_APARTMENT_NUMBER,
        MADDR_CITY,
        MADDR_STATE,
        MADDR_ZIP_CODE,
        PHONE,
        PARTY_AFFILIATION,
        VOTER_STATUS,
        WARD_NUMBER,
        PRECINCT_NUMBER,
        ACCOUNT_NUMBER,
        CENSUS_YEAR,
        CONFIDENCE,
    ],
    'Contracts':
    [
        CONTRACT,
        VENDOR_NUMBER,
        VENDOR,
    ],
    'CostHistory':
    [
        VENDOR_NUMBER,
        VENDOR_NAME,
        PHONE,
        EMAIL,
        ADDR_1,
        ADDR_2,
        CITY,
        STATE,
        ZIP,
    ],
    'ElectionHistory':
    [
    ],
    'ElectionModel_01':
    [
        RESIDENT_ID,
    ],
    'ElectionModel_02':
    [
        RESIDENT_ID,
    ],
    'ElectionModel_03':
    [
        RESIDENT_ID,
    ],
    'Employees':
    [
    ],
    'Gender_2014':
    [
    ],
    'GlcacJobsWithPermits_L':
    [
        JOB_NUMBER,
        PERMIT_NUMBER,
    ],
    'ParcelSummary_L':
    [
        IS_RESIDENTIAL,
        NORMALIZED_STREET_NAME,
        PARCEL_COUNT,
        BUILDING_COUNT,
        ELEC_ + BATHS_ + TOTAL,
        ELEC_ + KITCHENS_ + TOTAL,
        ELEC_ + OCCUPANCY_ + TOTAL,
        OIL_ + BATHS_ + TOTAL,
        OIL_ + KITCHENS_ + TOTAL,
        OIL_ + OCCUPANCY_ + TOTAL,
        GAS_ + BATHS_ + TOTAL,
        GAS_ + KITCHENS_ + TOTAL,
        GAS_ + OCCUPANCY_ + TOTAL,
        BATHS_ + TOTAL,
        KITCHENS_ + TOTAL,
        OCCUPANCY_ + TOTAL,
    ],
    'Parcels_L':
    [
        VISION_ID,
        ACCOUNT_NUMBER,
        MBLU,
        IS_RESIDENTIAL,
        OWNER_IS_LOCAL,
        NORMALIZED_STREET_NAME,
        LOCATION,
        YEAR_BUILT,
        AGE,
        BUILDING_COUNT,
        TOTAL_ASSESSED_VALUE,
        STYLE,
        HEATING_FUEL + _DESC,
        HEATING_TYPE + _DESC,
        AC_TYPE + _DESC,
        HEAT_AC,
        FIRST_FLOOR_USE,
        LAND_USE_CODE,
        LAND_USE_CODE + _DESC,
        RESIDENTIAL_UNITS,
        KITCHENS,
        BATHS,
        OCCUPANCY_HOUSEHOLDS,
        TOTAL_OCCUPANCY,
        TOTAL_BATHS,
        TOTAL_KITCHENS,
        TOTAL_AREA,
        LIVING_AREA,
        TOTAL_ACRES,
        SALE_PRICE,
        SALE_DATE,
        ZONE,
        OWNER_1_NAME,
        OWNER_2_NAME,
        OWNER_ADDRESS,
        OWNER_ZIP,
        NORMALIZED_ADDRESS,
        NORMALIZED_STREET_NUMBER,
        NORMALIZED_OCCUPANCY,
        NORMALIZED_ADDITIONAL_INFO,
    ],
    'LocalElectionResults':
    [
        ELECTION_DATE,
        OFFICE,
        CANDIDATE,
    ],
    'Partisans':
    [
        * COLUMN_GROUP['RESIDENT'],
        * COLUMN_GROUP['VOTER'],
        * COLUMN_GROUP['ADDRESS'],
        ZONING_CODE_1,
        PARTISAN_SCORE,
        RECENT_LOCAL_ELECTIONS_VOTED,
    ],
    'PollingPlaces':
    [
    ],
    'Precincts':
    [
    ],
    'RawElectricUsage':
    [
        YEAR,
        TOWN_NAME,
        SECTOR,
        JAN,
        FEB,
        MAR,
        APR,
        MAY,
        JUN,
        JUL,
        AUG,
        SEP,
        OCT,
        NOV,
        DEC,
        ANNUAL_ELECTRIC_USAGE,
    ],
    'RawGasUsage':
    [
        YEAR,
        TOWN_NAME,
        SECTOR,
        JAN,
        FEB,
        MAR,
        APR,
        MAY,
        JUN,
        JUL,
        AUG,
        SEP,
        OCT,
        NOV,
        DEC,
        ANNUAL_GAS_USAGE,
    ],
    'RawGeographicReport':
    [
        YEAR,
        COUNTY,
        TOWN_NAME,
        ZIP,
        SECTOR,
        ANNUAL_ELECTRIC_USAGE,
        ANNUAL_ELECTRIC_SAVINGS,
        ELECTRIC_INCENTIVES,
        ANNUAL_GAS_USAGE,
        ANNUAL_GAS_SAVINGS,
        GAS_INCENTIVES,
    ],
    'RawLocalElectionResults':
    [
        ELECTION_DATE,
        OFFICE_OR_CANDIDATE,
    ],
    'Residents':
    [
        * COLUMN_GROUP['RESIDENT'],
        * COLUMN_GROUP['SCORE'],
        * COLUMN_GROUP['VOTER'],
        * COLUMN_GROUP['HOME'],
        * COLUMN_GROUP['ADDRESS'],
        PARCEL_ID,
        ZONING_CODE_1,
    ],
    'Solar':
    [
    ],
    'Streets':
    [
    ],
    'Vision_Clean':
    [
        VISION_ID,
        ACCOUNT_NUMBER,
        MBLU,
        IS_RESIDENTIAL,
        OWNER_IS_LOCAL,
        LOCATION,
        YEAR_BUILT,
        AGE,
        BUILDING_COUNT,
        TOTAL_ASSESSED_VALUE,
        STYLE,
        HEATING_FUEL + _DESC,
        HEATING_TYPE + _DESC,
        AC_TYPE + _DESC,
        HEAT_AC,
        FIRST_FLOOR_USE,
        LAND_USE_CODE,
        LAND_USE_CODE + _DESC,
        RESIDENTIAL_UNITS,
        KITCHENS,
        BATHS,
        OCCUPANCY_HOUSEHOLDS,
        TOTAL_OCCUPANCY,
        TOTAL_BATHS,
        TOTAL_KITCHENS,
        TOTAL_AREA,
        LIVING_AREA,
        TOTAL_ACRES,
        SALE_PRICE,
        SALE_DATE,
        ZONE,
        OWNER_1_NAME,
        OWNER_2_NAME,
        OWNER_ADDRESS,
        OWNER_ZIP,
    ],
    'Water':
    [
        SERVICE_ID,
        LAST_NAME,
        FIRST_NAME,
        ADDR_STREET_NUMBER,
        ADDR_STREET_NAME,
        SERVICE_TYPE,
        DESCRIPTION,
        SERVICE,
        CURRENT_DATE,
        PRIOR_DATE,
        CURRENT_READING,
        PRIOR_READING,
    ],
}

COLUMN_ORDER['Partisans_' + D] = COLUMN_ORDER['Partisans']
COLUMN_ORDER['Partisans_' + R] = COLUMN_ORDER['Partisans']
COLUMN_ORDER['BuildingPermits_L_Solar_Summary'] = COLUMN_ORDER['BuildingPermits_L_Solar']
COLUMN_ORDER['Assessment_L_Residential'] = COLUMN_ORDER['Assessment_L_Commercial']

COLUMN_ORDER_TRAILING = \
{
    'BuildingPermits_L_Electrical':
    [
        * COLUMN_GROUP['PERMIT_DATES'],
    ],
    'BuildingPermits_L_Gas':
    [
        * COLUMN_GROUP['PERMIT_DATES'],
    ],
    'BuildingPermits_L_Plumbing':
    [
        * COLUMN_GROUP['PERMIT_DATES'],
    ],
    'BuildingPermits_L_Roof':
    [
        * COLUMN_GROUP['PERMIT_DATES_1'],
    ],
    'BuildingPermits_L_Siding':
    [
        * COLUMN_GROUP['PERMIT_DATES_1'],
    ],
    'BuildingPermits_L_Solar':
    [
        * COLUMN_GROUP['PERMIT_DATES'],
    ],
    'BuildingPermits_L_Solar_Summary':
    [
        * COLUMN_GROUP['PERMIT_DATES'],
    ],
    'BuildingPermits_L_Wx':
    [
        * COLUMN_GROUP['PERMIT_DATES_1'],
    ],
}


# Information on how to publish databases
PUBLISH_INFO = \
{
    'student': \
    {
        'number_columns': True,
        'drop_table_names':
        [
            'Employees',
            'Partisans_' + D,
            'Partisans_' + R,
        ],
        'encipher_column_names':
        [
            RESIDENT_ID
        ],
        'drop_column_names':
        [
            FIRST_NAME,
            MIDDLE_NAME,
            LAST_NAME,
            DATE_OF_BIRTH,
            OWNER_1_NAME,
            OWNER_2_NAME,
            OWNER_3_NAME,
            PHONE,
            GRANTOR,
            PREVIOUS_GRANTOR,
            PARTY_AFFILIATION,
            PARTY_PREFERENCE_SCORE,
            LEGACY_PREFERENCE_SCORE,
            GENDER,
            IS_FAMILY,
            LIKELY_DEM_SCORE,
            LEGACY_DEM_SCORE,
            D,
            R,
            ABSENT,
            VOTED_BOTH,
            CHANGED_AFFILIATION,
            PARTY_VOTED_HISTORY,
            LIKELY_DEM_COUNT,
            LIKELY_REPUB_COUNT,
            LOCAL_DEM_VOTER_COUNT,
            LOCAL_REPUB_VOTER_COUNT,
            MEAN_LIKELY_DEM_SCORE,
            MEAN_PARTY_PREFERENCE_SCORE,
            MEAN_LIKELY_DEM_VOTER_ENGAGEMENT_SCORE,
            MEAN_LIKELY_REPUB_VOTER_ENGAGEMENT_SCORE,
        ]
    },
    'town': \
    {
        'number_columns': True,
        'drop_table_names':
        [
            'Employees',
            'Gender_2014',
            'Lookup',
            'Partisans_' + D,
            'Partisans_' + R,
        ],
        'encipher_column_names':
        [
        ],
        'drop_column_names':
        [
            PARTY_AFFILIATION,
            PARTY_PREFERENCE_SCORE,
            LEGACY_PREFERENCE_SCORE,
            GENDER,
            IS_FAMILY,
            LIKELY_DEM_SCORE,
            LEGACY_DEM_SCORE,
            D,
            R,
            ABSENT,
            VOTED_BOTH,
            CHANGED_AFFILIATION,
            PARTY_VOTED_HISTORY,
            LIKELY_DEM_COUNT,
            LIKELY_REPUB_COUNT,
            LOCAL_DEM_VOTER_COUNT,
            LOCAL_REPUB_VOTER_COUNT,
            MEAN_LIKELY_DEM_SCORE,
            MEAN_PARTY_PREFERENCE_SCORE,
            MEAN_LIKELY_DEM_VOTER_ENGAGEMENT_SCORE,
            MEAN_LIKELY_REPUB_VOTER_ENGAGEMENT_SCORE,
        ]
    },
    'research': \
    {
        'number_columns': True,
        'drop_table_names':
        [
        ],
        'encipher_column_names':
        [
        ],
        'drop_column_names':
        [
        ]
    },
}


# Combine two dataframes by concatenating common columns and merging unique columns
def combine_dataframes( df_1, df_2, subset, keep, on ):

    common = df_1.columns.intersection( df_2.columns )
    diff_1_2 = df_1.columns.difference( df_2.columns )
    diff_2_1 = df_2.columns.difference( df_1.columns )

    # Concatenate common columns of two dataframes
    df_1_common = df_1[common]
    df_2_common = df_2[common]
    df_result = pd.concat( [df_1_common, df_2_common], ignore_index=True )
    df_result = df_result.drop_duplicates( subset=subset, keep=keep )

    # Merge in columns unique to df_1
    columns_to_merge = on + list( diff_1_2 )
    df_result = pd.merge( df_result, df_1[columns_to_merge], how='left', on=on )

    # Merge in columns unique to df_2
    columns_to_merge = on + list( diff_2_1 )
    df_result = pd.merge( df_result, df_2[columns_to_merge], how='left', on=on )

    return df_result


# Return column with no zero values, or all zero values
def nonzero( df, column ):

    # Drop zeroes from the column
    df_drop = df[ df[column]!=0 ]

    # If result is empty, use original dataframe
    if len( df_drop ) == 0:
        df_drop = df

    # Return the column
    return df_drop[column]


# Convert likely_dem_score to party_preference_score
def likely_dem_to_party_preference_score( dem_score ):

    if pd.isnull( dem_score ):
        pref_score = ''

    else:

        # Determine party prefix
        pref_score = ( 'D' if ( dem_score > 0 ) else ( 'R' if ( dem_score < 0 ) else '-' ) )

        # Append numeric score
        pref_score += ' ' + str( int( abs( dem_score ) ) ).zfill( 3 )

    return pref_score


# Isolate entries that did not find matches when merged with assessment data
def isolate_unmatched( df_merge, left_columns, df_result, s_descr='' ):

    # Create dataframe of unmatched entries
    df_unmatched = df_merge.copy()
    df_unmatched = df_unmatched[ df_unmatched[ACCOUNT_NUMBER].isna() ]
    df_unmatched = df_unmatched[left_columns]

    # Clear unmatched entries out of merged data
    df_matched = df_merge.copy()
    df_matched = df_matched.dropna( subset=[ACCOUNT_NUMBER] )

    # Append matched entries to the result
    df_result = df_result.append( df_matched, ignore_index=True )

    # Report progress
    print( '---' )
    if s_descr:
        print( '{}:'.format( s_descr ) )
    print( 'Matched: {}, Unmatched: {}'.format( df_matched.shape, df_unmatched.shape ) )
    print( 'Result: {}'.format( df_result.shape ) )

    return df_result, df_unmatched


# Add row to expanded address range
def add_address_row( address_number, street, row, df_expanded ):
    new_address = str( address_number ) + ' ' + street
    new_row = row.copy()
    new_row[NORMALIZED_ADDRESS] = new_address
    df_expanded = df_expanded.append( new_row, ignore_index=True )
    return( df_expanded )


# Expand dataframe such that each address range entry is replaced by a series of entries representing the range
def expand_address_ranges( df ):

    # Extract entries that represent implied address ranges, expressed as '<number>[letter]-<number>[letter]'
    df_ranges = df.copy()
    df_ranges = df_ranges[ df_ranges[NORMALIZED_ADDRESS].str.match( '^\d+[A-Z]*-\d+[A-Z]* .*$' ) ]

    # Generate a new dataframe that expands the ranges into individual addresses
    df_expanded = pd.DataFrame( columns=df_ranges.columns )

    for index, row in df_ranges.iterrows():

        # Extract numeric address range
        address_range = row[NORMALIZED_ADDRESS].split()[0].split( '-' )
        range_first = int( re.search( '^\d*', address_range[0] ).group(0) )
        range_last = int( re.search( '^\d*', address_range[1] ).group(0) )

        # Extract street
        words = row[NORMALIZED_ADDRESS].split()[1:]
        num_suffix = words.pop() if ( len( words ) and re.search( '^[A-Z]$', words[-1] ) ) else ''
        street = ' '.join( words )

        # Iterate over all numbers in the address range, incrementing by 2
        for num in range( range_first, range_last + 1, 2 ):
            df_expanded = add_address_row( str( num ) + num_suffix, street, row, df_expanded )

        # Determine whether either range specifier contains a letter
        letter_in_range_first = re.search( '[A-Z]', address_range[0] )
        letter_in_range_last = re.search( '[A-Z]', address_range[1] )

        # Include row representing second of two consecutive integer range specifiers
        if ( range_last == range_first + 1 ) and not ( letter_in_range_first or letter_in_range_last ):
            df_expanded = add_address_row( range_last, street, row, df_expanded )

        # Include literal range specifier containing letter, but only if numbers are equal
        if ( range_first == range_last ):
            if letter_in_range_first:
                df_expanded = add_address_row( address_range[0], street, row, df_expanded )
            if letter_in_range_last:
                df_expanded = add_address_row( address_range[1], street, row, df_expanded )

    # Extract entries that represent explicit address ranges, expressed as '<number>[letter]-<number>[letter]-<number>[letter]...'
    df_ranges = df.copy()
    df_ranges = df_ranges[ df_ranges[NORMALIZED_ADDRESS].str.match( '^\d+[A-Z]*-\d+[A-Z]*(-\d+[A-Z]*)+ .*$' ) ]

    # Iterate over explicit attress ranges
    for index, row in df_ranges.iterrows():

        # Extract street
        street = ' '.join( row[NORMALIZED_ADDRESS].split()[1:] )

        # Extract range parts
        address_range = row[NORMALIZED_ADDRESS].split()[0].split( '-' )

        # Iterate over range parts, adding a row for each
        for range_part in address_range:
            df_expanded = add_address_row( range_part, street, row, df_expanded )

    # Extract entries that represent ranges of occupancy identifiers, such as '95 A-B NEWTON ST'
    df_ranges = df.copy()
    df_ranges = df_ranges[ df_ranges[NORMALIZED_ADDRESS].str.match( '^\d+ [A-Z]-[A-Z] .*$' ) ]

    # Iterate over addresses with alphabetical ranges
    for index, row in df_ranges.iterrows():

        # Extract parts: address number, alphabetical range, and street name
        parts = row[NORMALIZED_ADDRESS].split()
        number = parts[0]
        alpha_range = parts[1].split( '-' )
        street = ' '.join( parts[2:] )

        # Construct list of letters in range
        letters_in_range = [chr(i) for i in range( ord( alpha_range[0] ), ord( alpha_range[1] ) + 1 )]

        # Iterate over letters, adding two entries for each: '<number> <letter>' and '<number><letter>'
        for letter in letters_in_range:
            address_number = ''.join( [number, letter] )
            df_expanded = add_address_row( address_number, street, row, df_expanded )

    df_no_ranges = df.loc[ df.index.difference( df_ranges.index ) ]
    df_expanded = df_expanded.append( df_no_ranges, ignore_index=True )

    return( df_expanded )


# Merge repeatedly on original and reformatted addresses
def merge_expand_merge_expand_merge( df_result, df_unmatched, left_columns, df_parcels, strip_left=False, strip_right=False ):

    # Merge unmatched with parcels table
    df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
    df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Unexpanded' )

    # Expand addresses in parcels table
    df_parcels = expand_address_ranges( df_parcels )

    # Merge unmatched with parcels table
    df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
    df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Expanded on right' )

    # Expand addresses in unmatched dataframe
    df_unmatched = expand_address_ranges( df_unmatched )

    # Merge unmatched with parcels table
    df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
    df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Expanded on left' )

    if strip_left:

        print( '---' )
        print( '-- Strip left --' )

        pattern = r'(^\d+)([A-Z]+ )(.*)'

        # Create copies of normalized address columns with trailing address letter stripped away
        df_unmatched[LEFT_ADDR_EDIT] = df_unmatched[NORMALIZED_ADDRESS].replace( { pattern : r'\1 \3 \2' }, regex=True )
        df_unmatched[LEFT_ADDR_STRIP] = df_unmatched[NORMALIZED_ADDRESS].replace( { pattern : r'\1 \3' }, regex=True )

        # Save columns
        left_columns = df_unmatched.columns

        # Merge unmatched edited
        df_unmatched[NORMALIZED_ADDRESS] = df_unmatched[LEFT_ADDR_EDIT]
        df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
        df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Edited on left' )

        # Here for the first time we remove meaningful information from the address.  Lower confidence level of subsequent matches.
        df_parcels[CONFIDENCE] = CONFIDENCE_MEDIUM

        # Merge unmatched stripped
        df_unmatched[NORMALIZED_ADDRESS] = df_unmatched[LEFT_ADDR_STRIP]
        df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
        df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Stripped on left' )

        # Delete columns no longer needed
        df_unmatched = df_unmatched.drop( columns=[LEFT_ADDR_EDIT, LEFT_ADDR_STRIP] )
        df_result = df_result.drop( columns=[LEFT_ADDR_EDIT, LEFT_ADDR_STRIP] )

    if strip_right:

        print( '---' )
        print( '-- Strip right --' )

        pattern = r'(^\d+)([A-Z]+ )(.*)'

        # Create copies of normalized address columns with trailing address letter stripped away
        df_parcels[RIGHT_ADDR_EDIT] = df_parcels[NORMALIZED_ADDRESS].replace( { pattern : r'\1 \3 \2' }, regex=True )
        df_parcels[RIGHT_ADDR_STRIP] = df_parcels[NORMALIZED_ADDRESS].replace( { pattern : r'\1 \3' }, regex=True )

        # Merge unmatched edited
        df_parcels[NORMALIZED_ADDRESS] = df_parcels[RIGHT_ADDR_EDIT]
        df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
        df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Edited on right' )

        # Merge unmatched stripped
        df_parcels[NORMALIZED_ADDRESS] = df_parcels[RIGHT_ADDR_STRIP]
        df_merge = pd.merge( df_unmatched, df_parcels, how='left', on=[NORMALIZED_ADDRESS] )
        df_result, df_unmatched = isolate_unmatched( df_merge, left_columns, df_result, 'Stripped on right' )

        # Delete columns no longer needed
        df_result = df_result.drop( columns=[RIGHT_ADDR_EDIT, RIGHT_ADDR_STRIP] )

    return df_result, df_unmatched


# Read parcels assessment table columns needed for merge
def read_parcels_table_for_merge( engine, columns=[NORMALIZED_ADDRESS, ACCOUNT_NUMBER, NORMALIZED_STREET_NUMBER, NORMALIZED_STREET_NAME] ):
    df_parcels = pd.read_sql_table( 'Parcels_L', engine, index_col=ID, columns=columns, parse_dates=True )
    df_parcels[RIGHT_ADDR_TRUNC] = df_parcels[NORMALIZED_STREET_NUMBER] + ' ' + df_parcels[NORMALIZED_STREET_NAME]
    df_parcels = df_parcels.drop( columns=[NORMALIZED_STREET_NUMBER, NORMALIZED_STREET_NAME] )
    return df_parcels


# Merge dataframe with commercial and residential assessment data based on normalized addresses
def merge_with_assessment_data( table_name, df_left, sort_by=[PERMIT_NUMBER, ACCOUNT_NUMBER], drop_subset=None, engine=None, df_parcels=None ):

    # If we have engine, retrieve the parcels table
    if engine != None:
        df_parcels = read_parcels_table_for_merge( engine )
    # else caller must supply the dataframe

    print( '---' )
    print( 'Left dataframe before merge: {}'.format( df_left.shape ) )

    # Save structures that we will need again later
    df_left[LEFT_ADDR_FULL] = df_left[NORMALIZED_ADDRESS]
    df_left[LEFT_ADDR_TRUNC] = df_left[NORMALIZED_STREET_NUMBER] + ' ' + df_left[NORMALIZED_STREET_NAME]

    # Initialize empty result and table of unmatched rows
    df_result = pd.DataFrame()
    left_columns = df_left.columns
    df_unmatched = df_left.copy()

    #
    # High confidence matching
    #

    # Initialize high confidence for these matches
    df_parcels[CONFIDENCE] = CONFIDENCE_HIGH

    # Match using full normalized address
    print( '---' )
    print( '-- Matching on left full and right full addresses (FxF) --' )
    df_result, df_unmatched = merge_expand_merge_expand_merge( df_result, df_unmatched, left_columns, df_parcels )

    # Retry using truncated address (street number + street name) on left side
    print( '---' )
    print( '-- Matching on left truncated and right full address (TxF) --' )
    df_unmatched[NORMALIZED_ADDRESS] = df_unmatched[LEFT_ADDR_TRUNC]
    df_result, df_unmatched = merge_expand_merge_expand_merge( df_result, df_unmatched, left_columns, df_parcels, strip_left=True )

    #
    # Low confidence matching
    #

    # Set low confidence for remaining matches
    df_parcels[CONFIDENCE] = CONFIDENCE_LOW

    # Getting more desperate.  Try again with truncated addresses on right side
    print( '---' )
    print( '-- Matching on left full and right truncated address (FxT) --' )
    df_unmatched[NORMALIZED_ADDRESS] = df_unmatched[LEFT_ADDR_FULL]
    df_parcels[NORMALIZED_ADDRESS] = df_parcels[RIGHT_ADDR_TRUNC]
    df_result, df_unmatched = merge_expand_merge_expand_merge( df_result, df_unmatched, left_columns, df_parcels, strip_right=True )

    # Finish up
    df_result = df_result.append( df_unmatched, ignore_index=True )
    df_result[NORMALIZED_ADDRESS] = df_result[LEFT_ADDR_FULL]
    df_result = df_result.drop( columns=[LEFT_ADDR_FULL, LEFT_ADDR_TRUNC, RIGHT_ADDR_TRUNC] )

    # When matches are duplicated with differing confidence levels, keep row with higher confidence
    df_result = df_result.sort_values( by=[CONFIDENCE], ascending=False )
    df_result = df_result.drop_duplicates( subset=set( df_result.columns ).difference( set( [CONFIDENCE] ) ) )

    # Clean up and sort according to supplied parameters
    df_result = df_result.drop_duplicates( subset=drop_subset )
    df_result = df_result.sort_values( by=sort_by )

    # Report final statistics
    len_unmatched = len( df_result[df_result[ACCOUNT_NUMBER].isna() ] )
    len_result = len( df_result )
    len_matched = len_result - len_unmatched
    print( '---' )
    print( "FINAL Matched in '{}': {}%".format( table_name, round( 100 * len_matched / len_result, 2 ) ) )
    print( 'FINAL Matched: {}'.format( len_matched ) )
    print( 'FINAL Unmatched: {}'.format( len_unmatched ) )
    print( 'FINAL Result: {}'.format( len_result ) )

    return df_result


# Read single input file with hyperlinks expanded
def read_excel_with_hyperlinks( input_filename, skiprows ):

    # Get the worksheet
    workbook = openpyxl.load_workbook( input_filename, data_only=True )
    worksheet = workbook.active

    # Determine index of column header row
    header_index = skiprows.stop if ( skiprows != None ) else 0

    # Extract column names and identify columns that contain hyperlinks
    col_names = []
    linked_cols = {}
    for row_index, row in enumerate( worksheet.iter_rows() ):
        if row_index == header_index:
            # Extract column names
            for cell in row:
                col_names.append( cell.value )
        elif row_index > header_index:
            # Determine which columns, if any, contain hyperlinks
            cell_index = -1
            for cell in row:
                cell_index += 1
                if cell.hyperlink != None:
                    linked_cols[col_names[cell_index]] = cell_index

    # Extract cell values and hyperlinks for dataframe
    data = []
    for row_index, row in enumerate( worksheet.iter_rows() ):

        if row_index > header_index:
            # It's a data row.  Load into the grid.
            data_row = []

            cell_index = -1
            for cell in row:
                cell_index += 1

                # Extract the value
                data_row.append( cell.value )

                # Optionally extract the hyperlink
                if col_names[cell_index] in linked_cols:
                    target = cell.hyperlink.target if cell.hyperlink != None else None
                    data_row.append( target )

            # Save row in grid
            data.append( data_row )

    # Build list of column names for dataframe
    columns = []
    for col_name in col_names:
        columns.append( col_name )
        if col_name in linked_cols:

            # Generate a unique column name
            n_suffix = 1
            linked_col_name = col_name + '_' + str( n_suffix )
            while linked_col_name in col_names:
                n_suffix += 1
                linked_col_name = col_name + n_suffix

            columns.append( linked_col_name )

    # Construct dataframe from extracted values
    df = pd.DataFrame( data=data, columns=columns )

    # Drop empty columns
    df = df.dropna( how='all', axis=1 )

    return df


# Read series of input files in specified directory
def read_excel_files( input_directory, column_labels, skiprows ):

    # Initialize empty dataframe
    df_xl = pd.DataFrame()

    # Construct dataframe from input files found in directory
    for filename in os.listdir( input_directory ):

        input_path = input_directory + '/' + filename
        print( 'Reading "{0}"'.format( input_path ) )

        # Filter warning apparently caused by unorthodox sheet name
        warnings.filterwarnings( 'ignore', category=UserWarning, module='openpyxl' )
        df = pd.read_excel( input_path, dtype=object, skiprows=skiprows )
        warnings.resetwarnings()

        # Optionally add extra columns and populate with filename fragments delimited by underscore - e.g., given input file moo_1234.xlsx, column value will be 'moo'
        if column_labels:
            labels = column_labels.split( ',' )
            values = filename.split( '_' )

            for label in labels:
                value = values.pop(0)
                df[label] = value
                print( '- Added column "{0}" with value "{1}"'.format( label, value ) )

        # Append partial input to dataframe
        df_xl = df_xl.append( df, sort=True )

    return df_xl


# Rename columns according to mappings
def rename_columns( df, table_name, exclude_unmapped=False, synthesize_unmapped=False ):

    # Strip whitespace from column names
    df.columns = df.columns.str.strip()

    # If any column names are not mapped, report and abort
    not_in = []
    if table_name in CONSISTENT_COLUMN_NAMES:
        for col_name in df.columns:
            if col_name not in CONSISTENT_COLUMN_NAMES[table_name]:
                not_in.append( col_name )
    else:
        not_in = list( df.columns )

    if len( not_in ):
        print( '!!! CONSISTENT_COLUMN_NAMES mappings missing for table {0}, {1} columns: {2}'.format( table_name, len(not_in), not_in ) )
        if exclude_unmapped:
            print( '!!! Excluding unmapped column names' )
            df = df.drop( columns=not_in )
            if len( df.columns ) == 0:
                print( '!!! No columns left in table after excluding unmapped column names' )
                exit()
        elif synthesize_unmapped:
            print( '!!! Synthesizing unmapped column names' )
            synthesize_unmapped_column_names( table_name, not_in )
        else:
            exit()

    # Correct misspelling and mislabeling
    if table_name in CONSISTENT_COLUMN_NAMES:

        bf = df.columns
        df = df.rename( index=str, columns=CONSISTENT_COLUMN_NAMES[table_name] )
        af = df.columns
        if len( bf.difference( af ) ):
            print( 'Renamed columns -' )
            print( '  From: {0}'.format( list( bf.difference( af ) ) ) )
            print( '    To: {0}'.format( list( af.difference( bf ) ) ) )

    return df


# Synthesize column names that do not have mappings
def synthesize_unmapped_column_names( table_name, not_in ):

    # Create entry if it doesn't exist
    if table_name not in CONSISTENT_COLUMN_NAMES:
        CONSISTENT_COLUMN_NAMES[table_name] = {}

    # Generate column name mappings
    for original_name in not_in:
        synthesized_name = "_".join( original_name.replace( '.', '_' ).replace( ':', '' ).replace( '%', 'pct' ).lower().split() )
        CONSISTENT_COLUMN_NAMES[table_name][original_name] = synthesized_name


# Open the SQLite database
def open_database( filename, b_create):

    print( '\nOpening database: ' + filename )

    # Optionally delete pre-existing database
    if b_create:
        if os.path.exists( filename ):
            print( ' Deleting...' )
            os.remove( filename )
        print( ' Creating...' )

    print( ' Connecting...' )
    conn = sqlite3.connect( filename )
    cur = conn.cursor()

    engine = sqlalchemy.create_engine( 'sqlite:///' + filename )

    return conn, cur, engine


def read_database( input_filename ):

    # Open the input database
    conn, cur, engine = open_database( input_filename, False )

    # Fetch names of all tables
    cur.execute( 'SELECT name FROM sqlite_master WHERE type="table" AND name NOT LIKE "sqlite_%";' )
    rows = cur.fetchall()

    # Built pandas representation of input database
    print( '' )
    input_db = {}
    for row in rows:
        table_name = row[0]
        print( 'Reading table', table_name )
        input_db[table_name] = pd.read_sql_table( table_name, engine, parse_dates=True )

    return input_db


# Create table with specified name and model
def create_table( table_name, conn, cur, columns=None, df=None, alt_column_order='' ):

    if ( columns is None ) and ( df is None ):
        print( "!!! create_table() missing required parameter: either 'columns' or 'df'" )
        exit()

    else:

        # Initialize and reorder columns
        if columns is None:
            columns = df.columns
        column_order = table_name if table_name in COLUMN_ORDER else alt_column_order
        columns = reorder_columns( column_order, list( columns ) )

        # Drop table if it already exists
        cur.execute( 'DROP TABLE IF EXISTS ' + table_name )

        # Generate SQL command
        create_sql = 'CREATE TABLE ' + table_name + ' ( id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE'
        for col_name in columns:
            sqltype = pdtype_to_sqltype( df, col_name )
            create_sql += ', "{0}" {1}'.format( col_name, sqltype )
        create_sql += ' )'

        # Execute the SQL command
        print( '' )
        print( create_sql )
        cur.execute( create_sql )

        # Optionally output dataframe to database and test spreadsheets
        if df is not None:
            df.to_sql( table_name, conn, if_exists='append', index=False )
            df = df.reindex( columns=columns )
            df.index = 1 + df.reset_index().index
            df.to_csv( '../test/' + table_name + '.csv', index_label=ID )

        conn.commit()


def reorder_columns( table_name, columns ):

    # Get leading and trailing column orders for this table
    leading_order = COLUMN_ORDER[table_name] if table_name in COLUMN_ORDER else []
    trailing_order = COLUMN_ORDER_TRAILING[table_name] if table_name in COLUMN_ORDER_TRAILING else []

    # Check for conflicts
    conflicts = list( set( leading_order ).intersection( set( trailing_order ) ) )
    if conflicts:
        print( '' )
        print( '!!! COLUMN_ORDER and COLUMN_ORDER_TRAILING for "{0}" contain {1} conflicts: "{2}"'.format( table_name, len( conflicts), conflicts ) )
        exit()

    # Create list of columns in preferred order
    any_order = [ x for x in columns if ( ( x not in leading_order ) and ( x not in trailing_order ) ) ]
    preferred_order = leading_order + any_order + trailing_order

    # Initialize list of remaining columns
    remaining = columns.copy()

    # Initialize empty result
    result = []

    # Iterate over column names listed in preferred order
    for col_name in preferred_order:

        if col_name in remaining:

            # Add to result
            result.append( col_name )

            # Remove from list of remaining remaining
            remaining.remove( col_name )

        else:
            print( '' )
            print( '!!! COLUMN_ORDER or COLUMN_ORDER_TRAILING for "{0}" table lists unexpected column name: "{1}"'.format( table_name, col_name ) )
            exit()

    # Append remaining remaining to result
    result += remaining

    # Correct misspelling and mislabeling
    if columns != result:
        print( '' )
        print( 'Reordered columns -' )
        print( '  From: {0}'.format( columns ) )
        print( '    To: {0}'.format( result ) )

    return result


def pdtype_to_sqltype( df, col_name ):

    if df is None:
        sqltype = 'TEXT'
    else:
        pdtype = str( df[col_name].dtype )

        if pdtype.startswith( 'float' ):
            sqltype = 'FLOAT'
        elif pdtype.startswith( 'int' ):
            sqltype = 'INT'
        else:
            sqltype = 'TEXT'

    return sqltype


# Prepare dataframe for saving to database
def prepare_for_database( df, table_name, exclude_unmapped=False, synthesize_unmapped=False ):

    # Correct misspelled and inconsistent column names
    df = rename_columns( df, table_name, exclude_unmapped, synthesize_unmapped )

    # Fix numeric columns
    df = fix_numeric_columns( df )

    # Fix date columns
    df = fix_date_columns( df )

    # Strip apostrophes
    df = strip_apostrophes( df )

    # Replace carriage returns
    df = replace_cr( df )

    # Collapse spaces
    df = collapse_spaces( df )

    # Strip leading and trailing spaces
    df = strip_ends( df )

    return df


# Remove apostrophes and apostrophe-like characters from name and address columns
def strip_apostrophes( df ):

    # Strip apostrophes from columns that follow strict conditions
    for column_name in df.columns:

        # Is this column a name or address fragment, but not a number?
        b_name = ( column_name.lower().find( 'name' ) != -1 )
        b_addr = ( column_name.lower().find( 'address' ) != -1 )
        b_num = ( column_name.lower().find( 'number' ) != -1 )

        if ( b_name or b_addr ) and not b_num:

            # Is the column type string?
            b_str = df[column_name].dtype == object

            if b_str:

                # Does this column have any apostrophes?
                sr1 = df[column_name].str.contains( "'" )
                sr2 = df[column_name].str.contains( "`" )
                counts1 = sr1.value_counts()
                counts2 = sr2.value_counts()
                b_has_apostrophe = ( True in counts1 ) or ( True in counts2 )

                # If the column has apostrophes...
                if b_has_apostrophe:

                    # Strip apostrophes from the column
                    df[column_name] = df[column_name].str.replace( r"['`]", '', regex=True )

    return df


# Replace all occurrences of carriage return character with single space
def replace_cr( df ):

    df = df.replace( to_replace=r'\r', value=' ', regex=True )

    return df


# Replace all occurrences of multiple consecutive spaces with single space
def collapse_spaces( df ):

    df = df.replace( to_replace=r' +', value=' ', regex=True )

    return df


# Strip leading and trailing spaces
def strip_ends( df ):

    df = df.replace( to_replace=r'^\s', value='', regex=True )
    df = df.replace( to_replace=r'\s$', value='', regex=True )

    return df


# Fix numeric columns in specified dataframe
def fix_numeric_columns( df ):

    for column_name in df.columns:

        if column_name.lower().find( 'zip' ) != -1:
            df[column_name] = fix_zip_code( df[column_name] )

        elif df[column_name].dtype == object:
            df[column_name] = pd.to_numeric( df[column_name], errors='ignore' )

    return df


# Reformat zip code when expressed as numeric value
def fix_zip_code( column ):

    # Build dictionary of replacement values
    dc_repl = {}
    for val in column.value_counts().index:

        if ( type( val ) in ( int, float ) ) or ( ( type( val ) == str ) and val.isdigit() ):
            # Convert numeric value to formatted zip code
            zip_plus_four = make_zip_code( int( val ) )

        else:
            # Retain original value
            zip_plus_four = val

        # Add formatted value to replacement dictionary
        dc_repl[val] = zip_plus_four

    # Replace original values with formatted zip codes
    column = column.replace( to_replace=dc_repl )

    return column


# Convert integer value to formatted zip code
def make_zip_code( i_val ):

    if i_val > 99999:
        # Value represents 9-digit zip
        i_four = i_val % 10000
        s_five = '{:05d}'.format( ( i_val - i_four ) // 10000 )
        s_four = '-{:04d}'.format( i_four )
    else:
        # Value represents 5-digit zip
        s_five = '{:05d}'.format( i_val )
        s_four = ''

    return '{0}{1}'.format( s_five, s_four )


# Force pandas to parse unparsed dates
def fix_date_columns( df ):

    # Determine which columns should represent dates
    expected_date_cols = get_named_date_columns( df )

    # Determine which columns represent dates
    date_cols = get_date_columns( df )

    fixed_cols = []

    # Fix columns that should be dates, but are not
    for col_name in expected_date_cols:

        if col_name not in date_cols:
            fixed_cols.append( col_name )
            df[col_name] = pd.to_datetime( df[col_name], infer_datetime_format=True, errors='coerce' )

    if len( fixed_cols ):
        print( '\nFixed date columns:', fixed_cols )

    return df


NOT_DATES = \
[
    OFFICE_OR_CANDIDATE,
]

# Find all columns whose names imply that they should represent dates
def get_named_date_columns( df ):
    date_cols = []
    for col_name in df.columns:
        if ( col_name.lower().find( 'date' ) != -1 ) and ( col_name not in NOT_DATES ):
            date_cols.append( col_name )
    return date_cols


# Find all columns that represent dates
def get_date_columns( df ):
    date_cols = []
    for col_name in df.columns:
        if df[col_name].dtype == 'datetime64[ns]':
            date_cols.append( col_name )
    return date_cols


# Encypher sensitive data
def encipher_column( df, column_name ):

    if column_name in df:
        df[column_name] = df[column_name].apply( encipher )

    return df


# Encipher alphanumeric characters in string
def encipher( s ):
    return transform( s, 1 ) if s else ''

# Decipher encipher() result
def decipher( s ):
    return transform( s, -1 ) if s else ''

TRANSFORM_SEQUENCE = string.digits + string.ascii_uppercase
TRANSFORM_SEQUENCE_LENGTH = len( TRANSFORM_SEQUENCE )
TRANSFORM_SHIFT = 13

def transform( text, direction ):

    trans = ''

    for c in text:

        offset = TRANSFORM_SEQUENCE.find( c )

        if offset == -1:
            trans += c
        else:
            trans += TRANSFORM_SEQUENCE[ ( offset + ( TRANSFORM_SHIFT * direction ) ) % TRANSFORM_SEQUENCE_LENGTH ]

    return trans


def make_df_about_energize_lawrence():
    copyright_notice = '© {} Energize Lawrence.  All rights reserved.'.format( datetime.date.today().year )
    df_about = pd.DataFrame( columns=['copyright'], data=[copyright_notice] )
    return df_about


def create_about_table( about_what, df_about, output_filename ):
    conn, cur, engine = open_database( output_filename, False )
    create_table( '_About' + about_what + 'Database', conn, cur, df=df_about )


def publish_database( input_db, output_filename, publish_info ):

    drop_table_names = publish_info['drop_table_names']
    drop_column_names = publish_info['drop_column_names']
    encipher_column_names = publish_info['encipher_column_names']
    number_columns = publish_info['number_columns']

    print( '' )
    print( 'Publishing database', output_filename )
    print( '- Dropping tables {0}'.format( drop_table_names ) )
    print( '- Dropping columns {0}'.format( drop_column_names ) )
    print( '- Enciphering columns {0}'.format( encipher_column_names ) )
    print( '' )

    output_db = {}
    for table_name in input_db:
        if table_name not in drop_table_names:
            print( 'Sanitizing table {0}'.format( table_name ) )
            output_db[table_name] = input_db[table_name].drop( columns=drop_column_names, errors='ignore' )
            for col_name in encipher_column_names:
                output_db[table_name] = encipher_column( output_db[table_name], col_name )

            if number_columns:
                n_cols = len( output_db[table_name].columns )
                num_width = len( str( n_cols ) )
                col_idx = 0
                for column_name in output_db[table_name].columns:
                    col_idx += 1
                    output_db[table_name] = output_db[table_name].rename( columns={ column_name: str( col_idx ).zfill( num_width ) + '-' + column_name } )


    # Open output database
    conn, cur, engine = open_database( output_filename, True )

    # Save result to database
    print( '' )
    for table_name in output_db:
        print( 'Publishing table', table_name )
        df = output_db[table_name]
        df.to_sql( table_name, conn, index=False )


def print_full( x ):
    pd.set_option( 'display.max_rows', len( x ) )
    print( x )
    pd.reset_option( 'display.max_rows' )



def report_elapsed_time( prefix='\n', start_time=START_TIME ):
    elapsed_time = round( ( time.time() - start_time ) * 1000 ) / 1000
    minutes, seconds = divmod( elapsed_time, 60 )
    ms = round( ( elapsed_time - int( elapsed_time ) ) * 1000 )
    print( prefix + 'Elapsed time: {:02d}:{:02d}.{:d}'.format( int( minutes ), int( seconds ), ms ) )
